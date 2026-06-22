"""
ping_monitor_gui.py — 핑감지 테스트기 v2
NC Agent 네트워크 종합 분석 도구
"""

import base64
import csv
import glob
import json
import math
import os
import re
import shutil
import struct
import subprocess
import sys
import threading
import time
import tkinter as tk
import zlib
from collections import deque
from datetime import datetime, date, timedelta
from tkinter import filedialog, messagebox, ttk

# 핵심 모듈 (core.py, startup.py)
try:
    from core import (SystemLogger, CsvLogger, FaultEngine, FaultPolicy,
                      DailyReporter, validate_config, validate_interval_warn)
    from startup import TrayManager, is_startup_enabled, set_startup
    _CORE_OK = True
except ImportError as _core_err:
    _CORE_OK = False

try:
    import psutil
    PSUTIL_OK = True
except ImportError:
    PSUTIL_OK = False

# ── 상수 ─────────────────────────────────────────────────────────────
NCAGENT_EXE = "NCAgent.exe"
# exe(PyInstaller) / .py 실행 모두 exe(또는 스크립트)와 같은 폴더 기준
_BASE_DIR    = (os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
                else os.path.dirname(os.path.abspath(__file__)))
CONFIG_FILE  = os.path.join(_BASE_DIR, "config.json")
_FLAG_FILE   = os.path.join(_BASE_DIR, ".running")   # 비정상 종료 감지용

H_PING   = ["DateTime", "Target", "Status", "ResponseTime_ms"]
H_PROC   = ["DateTime", "PID", "CPU_pct", "Memory_MB", "Event"]
H_NET    = ["DateTime", "Adapter", "Event", "IP", "Gateway"]
H_EVT    = ["DateTime", "Source", "EventID", "Level", "Message"]
H_FAULT  = ["DateTime", "Type", "Cause", "Detail"]
H_REPORT = ["Date", "Total_Ping", "Total_Fail", "Equip_Fail", "Server_Fail",
            "Max_RT_ms", "Avg_RT_ms", "Agent_Stop", "Simul_Fail"]

DEFAULT_TARGETS = [("설비 IP", "192.168.0.101"), ("서버", "hidc.cps.org")]

EVT_PROVIDERS = [
    "Tcpip", "DNS Client Events", "Service Control Manager",
    "Microsoft-Windows-Kernel-PnP", "Microsoft-Windows-NDIS",
    "e1000e", "RTL8153", "Realtek", "USB",
]

NC_SEARCH = [
    r"C:\Program Files\NCAgent", r"C:\Program Files (x86)\NCAgent",
    r"C:\NCAgent", r"D:\NCAgent",
    r"C:\HI-CPS",  r"C:\Program Files\HI-CPS", r"C:\Program Files (x86)\HI-CPS",
]

# ── 장애 원인 추정 매트릭스 ──────────────────────────────────────────
# (제목, 상황 설명, 권장 조치, 색상)
CAUSE_MATRIX = {
    "link_down": (
        "랜카드 / 케이블 연결 문제",
        "네트워크 어댑터 Link Down 감지\n"
        "→ PC가 네트워크에서 완전히 분리된 상태입니다.",
        "① 랜 케이블 연결 상태 확인\n"
        "② 스위치·허브 전원 및 포트 확인\n"
        "③ 네트워크 어댑터 드라이버 상태 확인",
        "#cc0000",
    ),
    "all_fail": (
        "로컬 네트워크 전체 장애",
        "설비 Ping 실패  +  서버 Ping 실패\n"
        "→ PC 자체의 네트워크 연결 또는 공용 스위치/라우터 문제 가능성이 높습니다.",
        "① 공유기·스위치 전원 재시작\n"
        "② PC 네트워크 설정 확인 (IP / 게이트웨이)\n"
        "③ 다른 장비에서 Ping 가능한지 교차 확인",
        "#cc0000",
    ),
    "equip_fail": (
        "사내 네트워크 또는 스위치 문제",
        "설비 Ping 실패  /  서버 Ping 정상\n"
        "→ 설비와 PC 사이의 내부 네트워크(스위치, 케이블, 설비 IP) 문제 가능성이 높습니다.",
        "① 설비 IP 주소 및 전원 상태 확인\n"
        "② 설비↔PC 구간 스위치 포트·케이블 확인\n"
        "③ 설비 측 네트워크 설정 확인",
        "#cc6600",
    ),
    "server_fail": (
        "서버 통신 문제 가능성 높음",
        "설비 Ping 정상  /  서버 Ping 실패\n"
        "→ 외부 서버(hidc.cps.org)와의 WAN/인터넷 통신 문제 가능성이 높습니다.",
        "① 인터넷(WAN) 연결 상태 확인\n"
        "② hidc.cps.org 서버 운영 상태 문의\n"
        "③ DNS 설정 및 방화벽 규칙 확인",
        "#cc6600",
    ),
    "agent_only": (
        "NC Agent 프로그램 문제",
        "Ping 정상  +  NC Agent 프로세스 종료\n"
        "→ NC Agent 프로그램 자체의 오류 또는 비정상 종료 가능성이 높습니다.",
        "① NC Agent 로그 파일 내용 확인\n"
        "② NC Agent 재시작\n"
        "③ Windows 이벤트 로그에서 오류 확인",
        "#cc6600",
    ),
    "normal": (
        "정상",
        "현재 감지된 장애 없음\n→ 모든 시스템이 정상 동작 중입니다.",
        "",
        "#006600",
    ),
}


# ── 헬퍼 ─────────────────────────────────────────────────────────────
_SCRIPT_DIR = _BASE_DIR   # CONFIG_FILE 정의 시 이미 계산됨


def _default_log_dir():
    """항상 스크립트 폴더 안의 logs 폴더를 기본 경로로 사용."""
    return os.path.join(_SCRIPT_DIR, "logs")


def _find_ncagent_dir():
    if PSUTIL_OK:
        try:
            for p in psutil.process_iter(["name", "exe"]):
                if p.info.get("name", "").lower() == NCAGENT_EXE.lower():
                    exe = p.info.get("exe", "")
                    if exe:
                        d = os.path.dirname(exe)
                        lg = os.path.join(d, "Log")
                        return lg if os.path.isdir(lg) else d
        except Exception:
            pass
    for base in NC_SEARCH:
        if not os.path.isdir(base):
            continue
        for sub in ("Log", "log", "Logs", ""):
            p = os.path.join(base, sub) if sub else base
            if os.path.isdir(p) and glob.glob(os.path.join(p, "*.log")):
                return p
        return base
    try:
        import winreg
        for hive in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
            for sub in (r"SOFTWARE\NCAgent", r"SOFTWARE\HI-CPS\NCAgent",
                        r"SOFTWARE\WOW6432Node\NCAgent"):
                try:
                    k = winreg.OpenKey(hive, sub)
                    path, _ = winreg.QueryValueEx(k, "InstallPath")
                    if os.path.isdir(path):
                        return path
                except Exception:
                    pass
    except Exception:
        pass
    return None


# ── 일별 통계 ─────────────────────────────────────────────────────────
class DayStats:
    def __init__(self):
        self.reset()

    def reset(self):
        self.date        = date.today()
        self.total_ping  = 0
        self.equip_fail  = 0
        self.server_fail = 0
        self.agent_stop  = 0
        self.net_error   = 0
        self.max_rt      = 0    # 최대 응답시간(ms)
        self.sum_rt      = 0    # 응답시간 합계 (평균 계산용)
        self.rt_count    = 0    # 응답 성공 횟수
        self.simul_fail  = 0    # 동시 장애 횟수 (전체 대상 동시 FAIL)

    @property
    def total_fail(self):
        return self.equip_fail + self.server_fail

    @property
    def avg_rt(self):
        return round(self.sum_rt / self.rt_count, 1) if self.rt_count > 0 else 0

    def snapshot(self):
        """현재 통계의 복사본 반환 (리셋 전 보고서 생성용)."""
        s = DayStats.__new__(DayStats)
        s.__dict__.update(self.__dict__)
        return s

    def to_row(self):
        return [str(self.date), self.total_ping, self.total_fail,
                self.equip_fail, self.server_fail,
                self.max_rt, self.avg_rt,
                self.agent_stop, self.simul_fail]


# ── 실시간 응답시간 그래프 ────────────────────────────────────────────
class RTGraph:
    """tkinter Canvas 기반 실시간 Ping 응답시간 그래프 (외부 라이브러리 불필요)."""

    MAXLEN = 720        # 1시간 @ 5초 간격
    C_BG   = "#1e1e2e"  # 배경
    C_GRID = "#2a2a3e"  # 그리드
    C_LINE = "#4fc3f7"  # 정상 선
    C_WARN = "#ffb74d"  # 급증 구간 (주황)
    C_FAIL = "#ef5350"  # FAIL 점
    C_MARK = "#ffd54f"  # 최대값 마커
    C_TEXT = "#90a4ae"  # 축 텍스트

    def __init__(self, parent, title):
        self._data = deque(maxlen=self.MAXLEN)  # (datetime, int|None)

        outer = ttk.LabelFrame(parent, text=title, padding=4)
        outer.pack(fill=tk.BOTH, expand=True)

        self.cv = tk.Canvas(outer, bg=self.C_BG,
                            highlightthickness=1, highlightbackground="#444")
        self.cv.pack(fill=tk.BOTH, expand=True)

        # 통계 바
        sf = ttk.Frame(outer)
        sf.pack(fill=tk.X, pady=(4, 0))
        self._lbl = {}
        for key, name in [("cur", "현재값"), ("max", "최대값"),
                           ("min", "최소값"), ("avg", "평균값")]:
            col = ttk.Frame(sf)
            col.pack(side=tk.LEFT, expand=True)
            ttk.Label(col, text=name, font=("", 8),
                      foreground="#888888").pack()
            v = ttk.Label(col, text="—", font=("", 9, "bold"))
            v.pack()
            self._lbl[key] = v

        self.cv.bind("<Motion>",    self._on_move)
        self.cv.bind("<Leave>",     self._on_leave)
        self.cv.bind("<Configure>", lambda e: self._draw())

    # ── 데이터 추가 (스레드 안전 — deque + GIL) ──────────────────────
    def push(self, rt):
        """rt: 응답시간(ms, int) 또는 None(FAIL)."""
        self._data.append((datetime.now(), rt))

    # ── 갱신 (반드시 메인 스레드에서 호출) ───────────────────────────
    def refresh(self):
        self._draw()
        self._update_stats()

    def _update_stats(self):
        vals = [r for _, r in self._data if r is not None]
        if not vals:
            for lb in self._lbl.values():
                lb.config(text="—")
            return
        self._lbl["cur"].config(text=f"{vals[-1]} ms")
        self._lbl["max"].config(text=f"{max(vals)} ms")
        self._lbl["min"].config(text=f"{min(vals)} ms")
        self._lbl["avg"].config(
            text=f"{round(sum(vals) / len(vals))} ms")

    def _margins(self):
        return 44, 8, 10, 24   # left, right, top, bottom

    def _draw(self):
        c = self.cv
        c.delete("graph")
        W, H = c.winfo_width(), c.winfo_height()
        if W < 30 or H < 30:
            return
        ML, MR, MT, MB = self._margins()
        gW, gH = W - ML - MR, H - MT - MB

        data = list(self._data)   # 스냅샷 (thread-safe with GIL)
        n    = len(data)
        vals = [r for _, r in data if r is not None]

        if not data:
            c.create_text(W // 2, H // 2, text="데이터 대기 중...",
                          fill=self.C_TEXT, font=("", 9), tags="graph")
            return

        if not vals:
            c.create_text(W // 2, H // 2, text="FAIL 연속 발생",
                          fill=self.C_FAIL, font=("", 9, "bold"), tags="graph")
            return

        y_max = max(max(vals) * 1.3, 10)

        # ── Y축 그리드 ───────────────────────────────────────
        for i in range(5):
            frac = i / 4
            yp   = MT + gH - int(gH * frac)
            c.create_line(ML, yp, ML + gW, yp,
                          fill=self.C_GRID, width=1, tags="graph")
            c.create_text(ML - 3, yp, text=f"{round(y_max * frac)}",
                          fill=self.C_TEXT, font=("", 7),
                          anchor=tk.E, tags="graph")

        # ── X축 시간 레이블 ───────────────────────────────────
        for frac, anc in [(0, tk.W), (0.5, tk.CENTER), (1, tk.E)]:
            idx = int(frac * (n - 1))
            xp  = ML + int(gW * frac)
            c.create_text(xp, MT + gH + 10,
                          text=data[idx][0].strftime("%H:%M:%S"),
                          fill=self.C_TEXT, font=("", 7),
                          anchor=anc, tags="graph")

        # ── 응답시간 선 + FAIL 점 ────────────────────────────
        prev_x = prev_y = prev_rt = None
        max_rt  = max(vals)
        max_idx = max(range(n), key=lambda i: data[i][1] or -1)

        for i, (ts, rt) in enumerate(data):
            xp = ML + int(gW * i / max(n - 1, 1))

            if rt is not None:
                yp = MT + gH - int(gH * rt / y_max)
                yp = max(MT, min(MT + gH, yp))

                if prev_x is not None and prev_rt is not None:
                    # 50% 이상 급증 → 주황색 강조
                    color = (self.C_WARN if rt > prev_rt * 1.5
                             else self.C_LINE)
                    c.create_line(prev_x, prev_y, xp, yp,
                                  fill=color, width=1.5,
                                  smooth=True, tags="graph")

                # 최대값 마커 (노란 원 + 텍스트)
                if i == max_idx:
                    c.create_oval(xp - 4, yp - 4, xp + 4, yp + 4,
                                  fill=self.C_MARK, outline="",
                                  tags="graph")
                    c.create_text(xp, yp - 10, text=f"{rt}ms",
                                  fill=self.C_MARK, font=("", 7, "bold"),
                                  tags="graph")

                prev_x, prev_y, prev_rt = xp, yp, rt

            else:
                # FAIL 빨간 점
                yp = MT + gH // 2
                c.create_oval(xp - 5, yp - 5, xp + 5, yp + 5,
                               fill=self.C_FAIL, outline="#ff8a80",
                               width=1, tags="graph")
                prev_x, prev_y, prev_rt = xp, yp, None

        # ── 축선 ─────────────────────────────────────────────
        c.create_line(ML, MT, ML, MT + gH + 1,
                      fill="#555555", tags="graph")
        c.create_line(ML, MT + gH, ML + gW, MT + gH,
                      fill="#555555", tags="graph")

    # ── 마우스 오버: 십자선 + 툴팁 ────────────────────────────────
    def _on_move(self, event):
        c = self.cv
        c.delete("hover")
        if not self._data:
            return
        W, H = c.winfo_width(), c.winfo_height()
        ML, MR, MT, MB = self._margins()
        gW = W - ML - MR

        x_rel = event.x - ML
        if x_rel < 0 or x_rel > gW:
            return

        data = list(self._data)
        n    = len(data)
        idx  = max(0, min(int(x_rel / max(gW, 1) * (n - 1)), n - 1))
        ts, rt = data[idx]
        xp = ML + int(gW * idx / max(n - 1, 1))

        # 수직 십자선
        c.create_line(xp, MT, xp, MT + H - MT - MB,
                      fill="#ffffff", width=1,
                      dash=(3, 3), tags="hover")

        # 툴팁 박스
        rt_text = f"{rt} ms" if rt is not None else "FAIL"
        tip     = f"{ts.strftime('%Y-%m-%d %H:%M:%S')}\n{rt_text}"
        tx = min(xp + 10, W - 100)
        ty = max(event.y - 38, MT + 2)
        c.create_rectangle(tx - 2, ty - 2, tx + 96, ty + 30,
                            fill="#2a2a4a", outline="#6666bb",
                            width=1, tags="hover")
        c.create_text(tx + 47, ty + 14, text=tip,
                      fill="#e0e8ff", font=("", 8),
                      justify=tk.CENTER, tags="hover")

    def _on_leave(self, event):
        self.cv.delete("hover")


# ── 대상 추가/수정 다이얼로그 ─────────────────────────────────────────
class TargetDialog(tk.Toplevel):
    def __init__(self, parent, name="", ip="", title="대상 추가"):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.result = None
        self.grab_set()
        self.transient(parent)

        ttk.Label(self, text="이름 (표시용):").grid(row=0, column=0, sticky=tk.W, padx=12, pady=8)
        self._name = tk.StringVar(value=name)
        ttk.Entry(self, textvariable=self._name, width=26).grid(row=0, column=1, padx=12, pady=8)

        ttk.Label(self, text="IP / 호스트명:").grid(row=1, column=0, sticky=tk.W, padx=12, pady=8)
        self._ip = tk.StringVar(value=ip)
        e = ttk.Entry(self, textvariable=self._ip, width=26)
        e.grid(row=1, column=1, padx=12, pady=8)
        e.focus_set()

        ttk.Separator(self, orient=tk.HORIZONTAL).grid(
            row=2, column=0, columnspan=2, sticky=tk.EW, padx=12)
        f = ttk.Frame(self)
        f.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(f, text="확인", command=self._ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(f, text="취소", command=self.destroy, width=10).pack(side=tk.LEFT, padx=5)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        ip = self._ip.get().strip()
        if not ip:
            messagebox.showwarning("입력 오류", "IP 또는 호스트명을 입력하세요.", parent=self)
            return
        self.result = (self._name.get().strip() or ip, ip)
        self.destroy()


# ── 메인 애플리케이션 ─────────────────────────────────────────────────
class PingMonitorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("핑감지 테스트기 v2")
        self.root.geometry("1060x920")
        self.root.minsize(860, 620)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        self._targets  = list(DEFAULT_TARGETS)
        self._running  = False
        self._threads  = []
        self._lock     = threading.Lock()
        self._interval       = tk.IntVar(value=5)
        self._log_dir        = tk.StringVar(value=_default_log_dir())
        self._retention_days = tk.IntVar(value=30)  # 로그 보관 기간(일)
        self._stats    = DayStats()

        # NC Agent 상태
        self._ncagent_dir        = None
        self._ncagent_pid        = None
        self._ncagent_was_up     = None
        self._ncagent_log_mtimes = {}

        # 네트워크 이전 상태
        self._prev_if_up  = {}
        self._prev_if_ips = {}
        self._prev_gw     = None

        # 이벤트 로그 마지막 수집 시간
        self._last_evt_collect = datetime.now()

        # 장애 지속시간 추적
        self._fail_start = {}    # host -> datetime (FAIL 시작 시각)
        self._prev_ok    = {}    # host -> bool     (직전 핑 결과)

        # 대시보드 위젯 참조
        self._dash = {}
        self._graph_equip  = None   # RTGraph — 설비 Ping
        self._graph_server = None   # RTGraph — 서버 Ping

        # ── 24시간 안정성 변수 ─────────────────────────────────────
        self._log_lock       = threading.Lock()  # 모든 CSV 쓰기 통합 잠금
        self._grace_until    = None              # 절전 복귀 유예 기간 종료 시각
        self._next_cycle_exp = None              # sleep 감지용 예상 시각
        self._last_ping_time = time.monotonic()  # watchdog 기준 시각
        self._last_save_time = time.monotonic()  # 마지막 config 저장 시각

        # core.py 모듈 인스턴스 (로거, 장애 엔진, 보고서)
        self._fault_policy = FaultPolicy() if _CORE_OK else None
        log_dir = _default_log_dir()
        if _CORE_OK:
            self._sys_log  = SystemLogger(log_dir)
            self._csv_log  = CsvLogger(log_dir, self._sys_log)
            self._reporter = DailyReporter(
                os.path.join(_SCRIPT_DIR, "reports"), self._sys_log)
        else:
            self._sys_log = self._csv_log = self._reporter = None
        self._engines = {}   # host -> FaultEngine (모니터링 시작 시 생성)

        # 트레이 아이콘
        if _CORE_OK:
            self._tray = TrayManager(
                on_show=self._tray_show,
                on_quit=self._tray_quit)
        else:
            self._tray = None
        self._minimize_to_tray = tk.BooleanVar(value=False)

        # 연속 실패 그룹 추적 (일시적 응답 누락 판단용)
        self._streak_count  = {}   # host -> 현재 연속 실패 수
        self._streak_start  = {}   # host -> 실패 시작 datetime
        self._streak_svr_ok = {}   # host -> 실패 시작 시 서버 Ping 상태
        self._streak_simul  = {}   # host -> 동시 실패 여부
        self._analysis_today = {   # 오늘 분석 집계
            "total": 0, "transient": 0, "fault": 0,
            "max_streak": 0, "max_dur": 0, "common": False,
        }

        self._build_ui()
        self._load_config()
        self._reload_target_tree()
        self._update_clock()
        self._watch_day()              # 모니터링 중지 상태에서도 자정 감지
        self._check_previous_crash()   # 이전 비정상 종료 여부 확인

        # 저장된 NC Agent 경로가 있으면 탭에 즉시 표시
        if self._ncagent_dir and os.path.isdir(self._ncagent_dir):
            self._ncagent_path_lbl.config(
                text=self._ncagent_dir, foreground="#333333")

    # ── 설정 저장/불러오기 ────────────────────────────────────────────
    def _save_config(self):
        # 저장 시: tuple(name, host) → dict 형식으로 변환
        targets_out = []
        for t in self._targets:
            if isinstance(t, (list, tuple)) and len(t) >= 2:
                name, host = str(t[0]), str(t[1])
                role = getattr(self, "_target_roles", {}).get(host, "equipment")
                targets_out.append({"name": name, "host": host, "role": role})
            elif isinstance(t, dict):
                targets_out.append(t)

        cfg = {
            "targets":        targets_out,
            "log_dir":        self._log_dir.get(),
            "report_dir":     os.path.join(_SCRIPT_DIR, "reports"),
            "interval":       self._interval.get(),
            "retention_days": self._retention_days.get(),
            "fault_policy": (self._fault_policy.as_dict()
                             if self._fault_policy else
                             {"suspect_fail_count": 3,
                              "fault_fail_count": 5,
                              "recovery_success_count": 3}),
            "last_date":    str(self._stats.date),
            "ncagent_dir":  self._ncagent_dir or "",  # 수동 지정 경로 유지
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_save_config", e)

    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_load_config", e, "config.json 읽기 실패")
            return

        # targets: 내부는 항상 tuple(name, host)로 유지 (기존 코드 호환)
        # role 정보는 FaultEngine 생성 시 config dict에서 별도 참조
        raw = d.get("targets")
        if isinstance(raw, list) and raw:
            parsed = []
            for t in raw:
                if isinstance(t, dict):
                    name = str(t.get("name", "")).strip()
                    host = str(t.get("host", "")).strip()
                    if host:
                        parsed.append((name, host))
                elif isinstance(t, (list, tuple)) and len(t) >= 2:
                    name = str(t[0]).strip()
                    host = str(t[1]).strip()
                    if host:
                        parsed.append((name, host))
            if parsed:
                self._targets = parsed

        # role 정보는 별도 보관 (FaultEngine용, index 기반 fallback)
        self._target_roles = {}
        if isinstance(raw, list):
            for i, t in enumerate(raw):
                if isinstance(t, dict):
                    host = str(t.get("host", "")).strip()
                    role = t.get("role", "equipment")
                    self._target_roles[host] = role

        saved_dir = d.get("log_dir", "").strip()
        if saved_dir:
            drive    = os.path.splitdrive(saved_dir)[0]
            drive_ok = (not drive) or os.path.exists(drive + os.sep)
            if drive_ok:
                self._log_dir.set(saved_dir)

        iv = d.get("interval")
        if isinstance(iv, int) and 1 <= iv <= 300:
            self._interval.set(iv)

        rd = d.get("retention_days")
        if isinstance(rd, int) and 1 <= rd <= 365:
            self._retention_days.set(rd)

        # fault_policy
        if _CORE_OK and d.get("fault_policy"):
            self._fault_policy = FaultPolicy(d)

        # 마지막 기록 날짜 → 재시작 시 날짜 바뀌었으면 롤오버 처리
        last_date_str = d.get("last_date", "")
        if last_date_str:
            try:
                last = date.fromisoformat(last_date_str)
                if last < date.today():
                    # 재시작 전에 날짜가 바뀜 → stats.date를 마지막 날짜로 설정
                    # _watch_day()가 1분 이내에 rollover 처리
                    self._stats.date = last
            except (ValueError, AttributeError):
                pass

        # 수동 지정된 NC Agent 경로 복원
        saved_nc = d.get("ncagent_dir", "").strip()
        if saved_nc and os.path.isdir(saved_nc):
            self._ncagent_dir = saved_nc

    # ── 로그 경로 프로퍼티 ────────────────────────────────────────────
    @property
    def _ld(self):           return self._log_dir.get()
    @property
    def _ping_path(self):    return os.path.join(self._ld, "ping_log.csv")
    @property
    def _fault_path(self):   return os.path.join(self._ld, "fail_log.csv")
    @property
    def _proc_path(self):    return os.path.join(self._ld, "process_log.csv")
    @property
    def _net_path(self):     return os.path.join(self._ld, "network_log.csv")
    @property
    def _evt_path(self):     return os.path.join(self._ld, "event_log.csv")
    @property
    def _report_path(self):  return os.path.join(self._ld, "daily_report.csv")

    def _ensure_logs(self):
        os.makedirs(self._ld, exist_ok=True)
        for path, hdr in [(self._ping_path,   H_PING),
                          (self._fault_path,  H_FAULT),
                          (self._proc_path,   H_PROC),
                          (self._net_path,    H_NET),
                          (self._evt_path,    H_EVT),
                          (self._report_path, H_REPORT)]:
            if not os.path.exists(path):
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    csv.writer(f).writerow(hdr)

    def _wcsv(self, path, row):
        """모든 CSV 쓰기를 _log_lock으로 보호."""
        with self._log_lock:
            try:
                with open(path, "a", newline="", encoding="utf-8-sig") as f:
                    csv.writer(f).writerow(row)
            except Exception as e:
                if self._sys_log:
                    self._sys_log.log("_wcsv", e, str(path))

    def _append_ui(self, tree, values, tag=None):
        try:
            kw = {"tags": (tag,)} if tag else {}
            tree.insert("", 0, values=values, **kw)
            kids = tree.get_children()
            if len(kids) > 2000:
                tree.delete(kids[-1])
        except tk.TclError:
            pass

    # ── UI 구성 ───────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_settings_strip()
        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8, pady=3)

        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 4))

        for label, builder in [
            ("  대시보드  ",      self._build_dashboard_tab),
            ("  Ping 로그  ",     self._build_ping_tab),
            ("  프로세스  ",      self._build_process_tab),
            ("  네트워크  ",      self._build_network_tab),
            ("  이벤트 로그  ",   self._build_event_tab),
            ("  NC Agent 로그  ", self._build_ncagent_log_tab),
            ("  장애 분석  ",     self._build_fault_tab),
        ]:
            f = ttk.Frame(nb)
            nb.add(f, text=label)
            builder(f)

        self._status_bar = ttk.Label(self.root, text="대기 중...",
                                      relief=tk.SUNKEN, anchor=tk.W)
        self._status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=8, pady=(2, 4))

    # ── 설정 스트립 ───────────────────────────────────────────────────
    def _build_settings_strip(self):
        frm = ttk.LabelFrame(self.root, text="설정", padding=6)
        frm.pack(fill=tk.X, padx=8, pady=(8, 0))

        left = ttk.Frame(frm)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # IP 목록 트리
        tf = ttk.Frame(left)
        tf.pack(fill=tk.X)
        self._target_tree = ttk.Treeview(tf, columns=("name", "ip"),
                                          show="headings", height=3, selectmode="browse")
        self._target_tree.heading("name", text="이름")
        self._target_tree.heading("ip",   text="IP / 호스트")
        self._target_tree.column("name",  width=90,  anchor=tk.W)
        self._target_tree.column("ip",    width=160, anchor=tk.W)
        self._target_tree.bind("<Double-1>", lambda e: self._edit_target())
        sb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=self._target_tree.yview)
        self._target_tree.configure(yscrollcommand=sb.set)
        self._target_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        # 버튼 행
        br = ttk.Frame(left)
        br.pack(fill=tk.X, pady=(4, 0))
        self._add_btn = ttk.Button(br, text="+ 추가", command=self._add_target, width=7)
        self._del_btn = ttk.Button(br, text="- 삭제", command=self._del_target, width=7)
        self._edt_btn = ttk.Button(br, text="수정",   command=self._edit_target, width=7)
        self._add_btn.pack(side=tk.LEFT, padx=2)
        self._del_btn.pack(side=tk.LEFT, padx=2)
        self._edt_btn.pack(side=tk.LEFT, padx=2)
        ttk.Label(br, text="  간격:").pack(side=tk.LEFT)
        ttk.Spinbox(br, from_=1, to=300, textvariable=self._interval, width=4).pack(side=tk.LEFT)
        ttk.Label(br, text="초").pack(side=tk.LEFT)

        # 로그 경로
        lr = ttk.Frame(left)
        lr.pack(fill=tk.X, pady=(4, 0))
        ttk.Label(lr, text="로그 경로:").pack(side=tk.LEFT)
        self._log_entry = ttk.Entry(lr, textvariable=self._log_dir)
        self._log_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        self._browse_btn = ttk.Button(lr, text="폴더 선택",
                                       command=self._browse_log_dir, width=9)
        self._browse_btn.pack(side=tk.LEFT)

        # 로그 보관 기간
        ar = ttk.Frame(left)
        ar.pack(fill=tk.X, pady=(4, 0))
        ttk.Label(ar, text="로그 보관:").pack(side=tk.LEFT)
        ttk.Spinbox(ar, from_=1, to=365, textvariable=self._retention_days,
                    width=4).pack(side=tk.LEFT, padx=2)
        ttk.Label(ar, text="일 경과 시 archive 폴더로 자동 이동").pack(side=tk.LEFT, padx=2)
        self._archive_btn = ttk.Button(ar, text="지금 정리",
                                        command=self._run_archive_now, width=9)
        self._archive_btn.pack(side=tk.LEFT, padx=(8, 0))
        self._archive_status = ttk.Label(ar, text="", foreground="#555555",
                                          font=("", 8))
        self._archive_status.pack(side=tk.LEFT, padx=4)

        # 시작/중지 + 자동 실행 버튼
        ctrl = ttk.Frame(frm)
        ctrl.pack(side=tk.RIGHT, fill=tk.Y, padx=(12, 0))
        self._start_btn = ttk.Button(ctrl, text="▶  시작", command=self._start, width=14)
        self._stop_btn  = ttk.Button(ctrl, text="■  중지", command=self._stop, width=14,
                                      state=tk.DISABLED)
        self._start_btn.pack(pady=4)
        self._stop_btn.pack(pady=4)
        _startup_text = ("자동 실행 해제" if (_CORE_OK and is_startup_enabled())
                         else "시작 시 자동 실행")
        self._startup_btn = ttk.Button(ctrl, text=_startup_text,
                                        command=self._toggle_startup, width=14)
        self._startup_btn.pack(pady=2)

    # ── 대시보드 탭 ───────────────────────────────────────────────────
    def _build_dashboard_tab(self, parent):
        # ── 상단 바: 현재 시간 + LED 인디케이터 ──────────────────────
        top = ttk.Frame(parent)
        top.pack(fill=tk.X, padx=12, pady=(8, 4))
        ttk.Label(top, text="현재 시간:", font=("", 10, "bold")).pack(side=tk.LEFT)
        self._dash["time"] = ttk.Label(top, text="-", font=("", 10))
        self._dash["time"].pack(side=tk.LEFT, padx=6)

        # Power / Status LED
        led_frm = ttk.LabelFrame(top, text="시스템", padding=4)
        led_frm.pack(side=tk.RIGHT, padx=6)
        # ttk 위젯은 background 직접 조회 불가 → 시스템 기본색 사용
        _bg = ttk.Style().lookup("TFrame", "background") or "#f0f0f0"
        for lbl, key, init_color in [
            ("Power",  "led_power",  "#888888"),
            ("Status", "led_status", "#888888"),
        ]:
            f = ttk.Frame(led_frm)
            f.pack(side=tk.LEFT, padx=10)
            cv = tk.Canvas(f, width=26, height=26,
                           highlightthickness=0, bg=_bg)
            cv.create_oval(3, 3, 23, 23, fill=init_color,
                           outline="#444444", width=1, tags="led")
            cv.pack()
            ttk.Label(f, text=lbl, font=("", 8, "bold")).pack()
            self._dash[key] = cv

        # 실시간 상태 카드 3개
        card_frm = ttk.LabelFrame(parent, text="실시간 상태", padding=8)
        card_frm.pack(fill=tk.X, padx=12, pady=4)

        card_defs = [
            ("설비 Ping",  "equip_status",  "equip_detail"),
            ("서버 Ping",  "server_status", "server_detail"),
            ("NC Agent",   "agent_status",  "agent_detail"),
        ]
        for i, (title, ks, kd) in enumerate(card_defs):
            card = ttk.LabelFrame(card_frm, text=title, padding=8)
            card.grid(row=0, column=i, padx=6, pady=2, sticky=tk.NSEW)
            card_frm.columnconfigure(i, weight=1)
            s = ttk.Label(card, text="-", font=("", 13, "bold"), width=12, anchor=tk.CENTER)
            s.pack()
            d = ttk.Label(card, text="-", font=("", 9), anchor=tk.CENTER)
            d.pack()
            self._dash[ks] = s
            self._dash[kd] = d

        # 네트워크 어댑터
        net = ttk.LabelFrame(parent, text="네트워크 어댑터", padding=8)
        net.pack(fill=tk.X, padx=12, pady=4)
        r1 = ttk.Frame(net)
        r1.pack(fill=tk.X)
        ttk.Label(r1, text="어댑터:").pack(side=tk.LEFT)
        self._dash["if_name"]   = ttk.Label(r1, text="-", width=32); self._dash["if_name"].pack(side=tk.LEFT, padx=4)
        self._dash["if_status"] = ttk.Label(r1, text="-", width=12); self._dash["if_status"].pack(side=tk.LEFT, padx=4)
        r2 = ttk.Frame(net)
        r2.pack(fill=tk.X, pady=2)
        ttk.Label(r2, text="IP:").pack(side=tk.LEFT)
        self._dash["if_ip"] = ttk.Label(r2, text="-", width=20); self._dash["if_ip"].pack(side=tk.LEFT, padx=4)
        ttk.Label(r2, text="  GW:").pack(side=tk.LEFT)
        self._dash["if_gw"] = ttk.Label(r2, text="-", width=16); self._dash["if_gw"].pack(side=tk.LEFT, padx=4)

        # 오늘 통계
        stat = ttk.LabelFrame(parent, text="오늘 통계", padding=8)
        stat.pack(fill=tk.X, padx=12, pady=4)
        stat_row = ttk.Frame(stat)
        stat_row.pack(fill=tk.X)
        for i, (lbl, key) in enumerate([("총 Ping", "s_ping"),
                                          ("설비 실패", "s_equip"),
                                          ("서버 실패", "s_server"),
                                          ("Agent 종료", "s_agent"),
                                          ("네트워크 오류", "s_net")]):
            col = ttk.Frame(stat_row)
            col.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=6)
            ttk.Label(col, text=lbl, font=("", 8), foreground="#666666").pack()
            v = ttk.Label(col, text="0", font=("", 18, "bold"), anchor=tk.CENTER)
            v.pack(fill=tk.X)
            self._dash[key] = v

        if not PSUTIL_OK:
            self._build_psutil_install_bar(parent)

        # ── 실시간 응답시간 그래프 ─────────────────────────────────────
        gp = tk.PanedWindow(parent, orient=tk.VERTICAL,
                             sashrelief=tk.RIDGE, sashwidth=6,
                             bg="#888888")
        gp.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 6))

        for attr, lbl in [("_graph_equip",  "설비 Ping 응답시간 (ms)"),
                           ("_graph_server", "서버 Ping 응답시간 (ms)")]:
            f = ttk.Frame(gp)
            gp.add(f, minsize=170)
            graph = RTGraph(f, lbl)
            setattr(self, attr, graph)

    # ── 로그 탭들 ────────────────────────────────────────────────────
    def _build_ping_tab(self, parent):
        self._ping_tree = self._make_tree(parent,
            ("dt","target","status","rt"),
            ("날짜/시간","대상","상태","응답시간(ms)"),
            (155, 220, 70, 110))
        self._add_save_bar(parent, self._ping_tree, "Ping 로그", "ping_log")

    def _build_process_tab(self, parent):
        ttk.Label(parent,
            text=f"  감시 대상: {NCAGENT_EXE}  |  5초마다 확인",
            foreground="#555555").pack(anchor=tk.W, pady=(6, 2))
        self._proc_tree = self._make_tree(parent,
            ("dt","pid","cpu","mem","event"),
            ("날짜/시간","PID","CPU(%)","메모리(MB)","이벤트"),
            (155, 80, 80, 100, 280))
        self._add_save_bar(parent, self._proc_tree, "프로세스 로그", "process_log")
        if not PSUTIL_OK:
            self._build_psutil_install_bar(parent)

    def _build_network_tab(self, parent):
        ttk.Label(parent,
            text="  Link Up/Down · IP 변경 · Gateway 변경 감지  |  5초마다 확인",
            foreground="#555555").pack(anchor=tk.W, pady=(6, 2))
        self._net_tree = self._make_tree(parent,
            ("dt","adapter","event","ip","gw"),
            ("날짜/시간","어댑터","이벤트","IP","Gateway"),
            (155, 160, 110, 140, 120))
        self._add_save_bar(parent, self._net_tree, "네트워크 로그", "network_log")
        if not PSUTIL_OK:
            self._build_psutil_install_bar(parent)

    def _build_event_tab(self, parent):
        ttk.Label(parent,
            text="  수집: Tcpip / DNS / SCM / Kernel-PnP / NDIS / Realtek / RTL8153  |  60초마다 수집",
            foreground="#555555").pack(anchor=tk.W, pady=(6, 2))
        self._evt_tree = self._make_tree(parent,
            ("dt","source","id","level","msg"),
            ("날짜/시간","소스","EventID","수준","메시지"),
            (155, 190, 70, 80, 380))
        self._add_save_bar(parent, self._evt_tree, "이벤트 로그", "event_log")

    def _build_ncagent_log_tab(self, parent):
        info = ttk.LabelFrame(parent, text="NC Agent 설치 경로", padding=6)
        info.pack(fill=tk.X, padx=8, pady=6)
        r = ttk.Frame(info)
        r.pack(fill=tk.X)
        ttk.Label(r, text="경로:").pack(side=tk.LEFT)
        self._ncagent_path_lbl = ttk.Label(r, text="모니터링 시작 후 자동 탐색",
                                            foreground="#888888")
        self._ncagent_path_lbl.pack(side=tk.LEFT, padx=6, fill=tk.X, expand=True)
        ttk.Button(r, text="직접 선택",
                   command=self._browse_ncagent_dir, width=10).pack(side=tk.RIGHT)

        self._nclog_tree = self._make_tree(parent,
            ("dt","fname","action"),
            ("날짜/시간","파일명","처리 내용"),
            (155, 260, 380))
        self._add_save_bar(parent, self._nclog_tree, "NC Agent 로그", "ncagent_log")

    def _build_fault_tab(self, parent):
        # ── 오늘 응답 누락 분석 요약 박스 ────────────────────────────
        sum_frm = ttk.LabelFrame(parent, text="오늘 응답 누락 분석 요약", padding=8)
        sum_frm.pack(fill=tk.X, padx=8, pady=(6, 4))

        for lbl, key, init in [
            ("총 실패 그룹",   "as_total",     "0"),
            ("일시적 누락",    "as_transient", "0"),
            ("실제 장애 의심", "as_fault",     "0"),
            ("최대 연속 실패", "as_max_streak","—"),
            ("최대 지속시간",  "as_max_dur",   "—"),
            ("공통 장애",      "as_common",    "없음"),
        ]:
            col = ttk.Frame(sum_frm)
            col.pack(side=tk.LEFT, expand=True, padx=4)
            ttk.Label(col, text=lbl, font=("", 8),
                      foreground="#666666").pack()
            v = ttk.Label(col, text=init, font=("", 13, "bold"),
                          anchor=tk.CENTER)
            v.pack(fill=tk.X)
            self._dash[key] = v

        # ── 서브 노트북: 원인 추정 & 장애 이력 / 응답 누락 분석 ──────
        sub = ttk.Notebook(parent)
        sub.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        t1 = ttk.Frame(sub)
        t2 = ttk.Frame(sub)
        sub.add(t1, text="  원인 추정 & 장애 이력  ")
        sub.add(t2, text="  응답 누락 분석  ")

        # ── 탭1: 기존 원인 추정 패널 + 장애 이력 ─────────────────────
        est = ttk.LabelFrame(t1, text="원인 추정", padding=10)
        est.pack(fill=tk.X, padx=8, pady=(6, 4))

        left = ttk.Frame(est)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 16))
        for key, label in [
            ("ce_equip",   "설비 Ping"),
            ("ce_server",  "서버 Ping"),
            ("ce_agent",   "NC Agent"),
            ("ce_network", "네트워크"),
        ]:
            row = ttk.Frame(left)
            row.pack(fill=tk.X, pady=3)
            ttk.Label(row, text=f"{label}:", width=11,
                      font=("", 9, "bold"), anchor=tk.W).pack(side=tk.LEFT)
            lbl = ttk.Label(row, text="—", width=28, font=("", 9))
            lbl.pack(side=tk.LEFT)
            self._dash[key] = lbl

        ttk.Separator(est, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        right = ttk.Frame(est)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._dash["ce_title"] = ttk.Label(
            right, text="모니터링 시작 후 분석됩니다.",
            font=("", 11, "bold"), foreground="#555555")
        self._dash["ce_title"].pack(anchor=tk.W)
        self._dash["ce_cause"] = ttk.Label(
            right, text="", font=("", 9), justify=tk.LEFT, foreground="#333333")
        self._dash["ce_cause"].pack(anchor=tk.W, padx=4, pady=(4, 2))
        ttk.Separator(right, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)
        ttk.Label(right, text="권장 조치:", font=("", 9, "bold"),
                  foreground="#444444").pack(anchor=tk.W)
        self._dash["ce_action"] = ttk.Label(
            right, text="", font=("", 9), justify=tk.LEFT, foreground="#555555")
        self._dash["ce_action"].pack(anchor=tk.W, padx=4, pady=(2, 0))

        ttk.Label(t1,
            text="  Ping 실패 / NCAgent 종료 / 네트워크 오류 / Windows 이벤트 오류 자동 기록",
            foreground="#555555").pack(anchor=tk.W, pady=(4, 2))
        self._fault_tree = self._make_tree(t1,
            ("dt","ftype","cause","detail"),
            ("발생시간","유형","원인","상세내용"),
            (155, 110, 200, 380))
        self._add_save_bar(t1, self._fault_tree, "장애 분석", "fault_log")

        # ── 탭2: 응답 누락 분석 테이블 ───────────────────────────────
        ttk.Label(t2,
            text="  연속 실패 횟수 기준으로 일시적 응답 누락 / 실제 장애를 자동 판별합니다."
                 "  ■ 노란색 = 일시적 누락  ■ 빨간색 = 실제 장애  ■ 보라색 = 공통 구간",
            foreground="#555555", wraplength=800, justify=tk.LEFT,
        ).pack(anchor=tk.W, padx=8, pady=(6, 2))

        self._analysis_tree = self._make_tree(t2,
            ("ts","name","streak","dur","svr","eq","verdict","cause","action"),
            ("분석시간","설비명","연속실패","지속시간","서버Ping","설비Ping",
             "판정","추정원인","권장조치"),
            (135, 80, 60, 70, 60, 65, 150, 200, 230))

        self._analysis_tree.tag_configure(
            "TRANSIENT", foreground="#b87a00")
        self._analysis_tree.tag_configure(
            "FAULT",     foreground="#cc0000", font=("", 9, "bold"))
        self._analysis_tree.tag_configure(
            "COMMON",    foreground="#7700aa", font=("", 9, "bold"))

        self._add_save_bar(t2, self._analysis_tree, "응답 누락 분석", "analysis_log")

    # ── 공통 트리/저장 ────────────────────────────────────────────────
    @staticmethod
    def _make_tree(parent, cols, headings, widths):
        f = ttk.Frame(parent)
        f.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)
        tree = ttk.Treeview(f, columns=cols, show="headings")
        for c, h, w in zip(cols, headings, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor=tk.W)
        tree.tag_configure("OK",   foreground="#006600")
        tree.tag_configure("FAIL", foreground="#cc0000", font=("", 9, "bold"))
        tree.tag_configure("WARN", foreground="#cc6600")
        sb = ttk.Scrollbar(f, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        return tree

    def _add_save_bar(self, parent, tree, label, prefix):
        bar = ttk.Frame(parent)
        bar.pack(fill=tk.X, padx=6, pady=(0, 2))
        ttk.Label(bar, text=f"화면에 표시된 {label}을(를) CSV로 저장합니다.",
                  foreground="#555555").pack(side=tk.LEFT, padx=4)
        ttk.Button(bar, text="📥  CSV로 저장",
                   command=lambda t=tree, l=label, p=prefix: self._save_csv(t, l, p),
                   width=16).pack(side=tk.RIGHT, padx=4)

    def _save_csv(self, tree, label, prefix):
        rows = [tree.item(iid)["values"] for iid in reversed(tree.get_children())]
        if not rows:
            messagebox.showinfo("알림", f"저장할 {label} 데이터가 없습니다.")
            return
        name = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path = filedialog.asksaveasfilename(title=f"{label} 저장",
            initialfile=name, defaultextension=".csv",
            filetypes=[("CSV 파일", "*.csv"), ("모든 파일", "*.*")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerows(rows)
            messagebox.showinfo("저장 완료", f"총 {len(rows)}건 저장되었습니다.\n\n{path}")
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))

    # ── 대상 관리 ─────────────────────────────────────────────────────
    @staticmethod
    def _t(entry) -> tuple:
        """target 항목(dict 또는 tuple)에서 (name, host, role) 반환."""
        if isinstance(entry, dict):
            return entry.get("name",""), entry.get("host",""), entry.get("role","equipment")
        if isinstance(entry, (list, tuple)):
            name = entry[0] if len(entry) > 0 else ""
            host = entry[1] if len(entry) > 1 else ""
            return name, host, "equipment"
        return str(entry), "", "equipment"

    def _reload_target_tree(self):
        self._target_tree.delete(*self._target_tree.get_children())
        for t in self._targets:
            name, host, _ = self._t(t)
            self._target_tree.insert("", tk.END, values=(name, host))

    def _add_target(self):
        dlg = TargetDialog(self.root)
        self.root.wait_window(dlg)
        if dlg.result:
            self._targets.append(dlg.result)
            self._reload_target_tree()

    def _del_target(self):
        sel = self._target_tree.selection()
        if not sel:
            return
        idx = self._target_tree.index(sel[0])
        t0 = self._targets[idx]
        t0_name = str(t0[0]) if isinstance(t0,(list,tuple)) else t0.get("name","")
        if messagebox.askyesno("삭제 확인", f"'{t0_name}' 삭제하시겠습니까?"):
            self._targets.pop(idx)
            self._reload_target_tree()

    def _edit_target(self):
        sel = self._target_tree.selection()
        if not sel:
            return
        idx = self._target_tree.index(sel[0])
        te  = self._targets[idx]
        name = str(te[0]) if isinstance(te,(list,tuple)) else te.get("name","")
        ip   = str(te[1]) if isinstance(te,(list,tuple)) and len(te)>1 else te.get("host","")
        dlg = TargetDialog(self.root, name=name, ip=ip, title="대상 수정")
        self.root.wait_window(dlg)
        if dlg.result:
            self._targets[idx] = dlg.result
            self._reload_target_tree()

    def _browse_log_dir(self):
        cur = self._log_dir.get()
        init = cur if os.path.exists(cur) else os.path.expanduser("~")
        d = filedialog.askdirectory(title="로그 저장 폴더 선택", initialdir=init)
        if d:
            self._log_dir.set(d)

    # ── 시계 ──────────────────────────────────────────────────────────
    def _update_clock(self):
        try:
            self._dash["time"].config(text=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            self.root.after(1000, self._update_clock)
        except tk.TclError:
            pass

    # ── Ping ──────────────────────────────────────────────────────────
    @staticmethod
    def _ping(host):
        try:
            r = subprocess.run(
                ["ping", "-n", "1", "-w", "2000", host],
                capture_output=True, text=True, timeout=6,
                creationflags=subprocess.CREATE_NO_WINDOW)
            if r.returncode == 0:
                m = re.search(r"(?:time|시간)[=<](\d+)ms", r.stdout)
                return True, int(m.group(1)) if m else 0
            return False, None
        except Exception:
            return False, None

    # ── 프로세스 감시 ─────────────────────────────────────────────────
    def _check_process(self):
        if not PSUTIL_OK:
            return
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        found = None
        try:
            for p in psutil.process_iter(["name", "pid", "cpu_percent", "memory_info"]):
                if p.info.get("name", "").lower() == NCAGENT_EXE.lower():
                    found = p
                    break
        except Exception:
            pass

        if found:
            try:
                pid = found.pid
                cpu = round(found.cpu_percent(interval=0.2), 1)
                mem = round(found.memory_info().rss / 1024 / 1024, 1)
                evt = "시작 감지" if self._ncagent_pid != pid and self._ncagent_pid is not None else (
                      "재시작" if self._ncagent_pid is not None and self._ncagent_pid != pid else "실행중")
                self._ncagent_pid    = pid
                self._ncagent_was_up = True
                row = [now, pid, cpu, mem, evt]
                self._wcsv(self._proc_path, row)
                tag = "WARN" if "감지" in evt else "OK"
                self.root.after(0, lambda r=row, t=tag, c=cpu, m=mem, p=pid, e=evt: (
                    self._append_ui(self._proc_tree, r, t),
                    self._dash["agent_status"].config(text="실행중", foreground="#006600"),
                    self._dash["agent_detail"].config(text=f"PID:{p}  CPU:{c}%  Mem:{m}MB"),
                ))
            except Exception:
                pass
        else:
            if self._ncagent_was_up:
                row = [now, "-", "-", "-", "종료 감지"]
                self._wcsv(self._proc_path, row)
                self._record_fault(now, "프로세스", f"{NCAGENT_EXE} 종료",
                                   "프로세스가 비정상 종료되었습니다.")
                with self._lock:
                    self._stats.agent_stop += 1
                self._ncagent_pid = None
                self.root.after(0, lambda r=row: (
                    self._append_ui(self._proc_tree, r, "FAIL"),
                    self._dash["agent_status"].config(text="종료됨", foreground="#cc0000"),
                    self._dash["agent_detail"].config(text="-"),
                ))
            elif self._ncagent_was_up is None:
                self.root.after(0, lambda: (
                    self._dash["agent_status"].config(text="미실행", foreground="#999999"),
                    self._dash["agent_detail"].config(text="-"),
                ))
            self._ncagent_was_up = False

    # ── 네트워크 어댑터 감시 ──────────────────────────────────────────
    def _check_network(self):
        if not PSUTIL_OK:
            return
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            stats = psutil.net_if_stats()
            addrs = psutil.net_if_addrs()
        except Exception:
            return

        gw = self._get_gateway()
        primary_name = primary_ip = ""

        for iface, stat in stats.items():
            if iface.lower() in ("lo", "loopback"):
                continue
            is_up = stat.isup
            prev  = self._prev_if_up.get(iface)

            if prev is not None and prev != is_up:
                event = "Link Up" if is_up else "Link Down"
                ips = ",".join(a.address for a in addrs.get(iface, []) if a.family == 2)
                row = [now, iface, event, ips, gw or ""]
                self._wcsv(self._net_path, row)
                tag = "OK" if is_up else "FAIL"
                self.root.after(0, lambda r=row, t=tag: self._append_ui(self._net_tree, r, t))
                if not is_up:
                    self._record_fault(now, "네트워크", f"어댑터 Link Down", f"{iface}")
                    with self._lock:
                        self._stats.net_error += 1

            self._prev_if_up[iface] = is_up

            cur_ips = {a.address for a in addrs.get(iface, []) if a.family == 2}
            prev_ips = self._prev_if_ips.get(iface)
            if prev_ips is not None and prev_ips != cur_ips:
                row = [now, iface, "IP 변경",
                       f"{','.join(prev_ips)} -> {','.join(cur_ips)}", gw or ""]
                self._wcsv(self._net_path, row)
                self.root.after(0, lambda r=row: self._append_ui(self._net_tree, r, "WARN"))
            self._prev_if_ips[iface] = cur_ips

            if is_up and cur_ips and not primary_name:
                primary_name = iface
                primary_ip   = ",".join(cur_ips)

        if self._prev_gw is not None and gw and self._prev_gw != gw:
            row = [now, "System", "Gateway 변경",
                   f"{self._prev_gw} -> {gw}", gw]
            self._wcsv(self._net_path, row)
            self.root.after(0, lambda r=row: self._append_ui(self._net_tree, r, "WARN"))
        self._prev_gw = gw

        self.root.after(0, lambda n=primary_name, ip=primary_ip, g=gw: (
            self._dash["if_name"].config(text=n or "-"),
            self._dash["if_ip"].config(text=ip or "-"),
            self._dash["if_gw"].config(text=g or "-"),
            self._dash["if_status"].config(
                text="Link Up" if n else "Link Down",
                foreground="#006600" if n else "#cc0000"),
        ))

    @staticmethod
    def _get_gateway():
        try:
            r = subprocess.run(["ipconfig"], capture_output=True,
                               timeout=4, creationflags=subprocess.CREATE_NO_WINDOW)
            for enc in ("cp949", "utf-8", "euc-kr"):
                try:
                    text = r.stdout.decode(enc)
                    m = re.search(r"(?:기본 게이트웨이|Default Gateway)[^:\d]*([\d.]+)", text)
                    if m:
                        return m.group(1)
                except Exception:
                    continue
        except Exception:
            pass
        return None

    # ── Windows 이벤트 로그 수집 ──────────────────────────────────────
    def _collect_events(self):
        minutes = max(2, self._interval.get() + 1)
        providers = ",".join(f"'{p}'" for p in EVT_PROVIDERS)
        ps = (
            f"$t=(Get-Date).AddMinutes(-{minutes});"
            f"$pv=@({providers});"
            "$ev=Get-WinEvent -FilterHashtable @{LogName='System';StartTime=$t} "
            "-ErrorAction SilentlyContinue | Where-Object {$pv -contains $_.ProviderName};"
            "if($ev){$ev | Select-Object -First 40 | ForEach-Object {"
            "$dt=$_.TimeCreated.ToString('yyyy-MM-dd HH:mm:ss');"
            "$src=$_.ProviderName;"
            "$id=$_.Id;"
            "$lv=$_.LevelDisplayName;"
            "$mg=(($_.Message -split \"`n\")[0] -replace '`r','').Substring("
            "0,[Math]::Min(200,(($_.Message -split \"`n\")[0]).Length));"
            "\"$dt`t$src`t$id`t$lv`t$mg\"}}"
        )
        try:
            r = subprocess.run(
                ["powershell", "-NonInteractive", "-NoProfile", "-Command", ps],
                capture_output=True, timeout=20,
                creationflags=subprocess.CREATE_NO_WINDOW)
            text = ""
            for enc in ("utf-8", "cp949", "utf-16"):
                try:
                    text = r.stdout.decode(enc)
                    break
                except Exception:
                    continue
            for line in text.strip().splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t", 4)
                if len(parts) < 4:
                    continue
                dt, src, eid, lvl = parts[:4]
                msg = parts[4] if len(parts) > 4 else ""
                row = [dt, src, eid, lvl, msg]
                self._wcsv(self._evt_path, row)
                tag = "FAIL" if lvl in ("오류", "Error") else (
                      "WARN" if lvl in ("경고", "Warning") else None)
                self.root.after(0, lambda r=row, t=tag: self._append_ui(self._evt_tree, r, t))
                if lvl in ("오류", "Error"):
                    self._record_fault(dt, "Windows 이벤트",
                                       f"[{src}] ID:{eid}", msg[:120])
                    with self._lock:
                        self._stats.net_error += 1
        except Exception:
            pass

    # ── NC Agent 로그 파일 감시 ───────────────────────────────────────
    def _check_ncagent_logs(self):
        if self._ncagent_dir is None:
            found = _find_ncagent_dir() or ""
            self._ncagent_dir = found
            self.root.after(0, lambda d=found: self._ncagent_path_lbl.config(
                text=d if d else "NCAgent 경로를 찾을 수 없습니다.",
                foreground="#333333" if d else "#cc0000"))

        if not self._ncagent_dir or not os.path.isdir(self._ncagent_dir):
            return

        backup = os.path.join(self._ld, "ncagent_backup")
        os.makedirs(backup, exist_ok=True)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for logf in glob.glob(os.path.join(self._ncagent_dir, "*.log")):
            fname = os.path.basename(logf)
            try:
                mtime = os.path.getmtime(logf)
            except Exception:
                continue
            prev = self._ncagent_log_mtimes.get(fname, 0)
            if mtime == prev:
                continue
            action = "발견" if prev == 0 else "업데이트"
            try:
                dst = os.path.join(backup,
                                   f"{date.today().strftime('%Y%m%d')}_{fname}")
                shutil.copy2(logf, dst)
                action += " → 백업 완료"
            except Exception as e:
                action += f" → 백업 실패: {e}"
            self._ncagent_log_mtimes[fname] = mtime
            row = [now, fname, action]
            self.root.after(0, lambda r=row: self._append_ui(self._nclog_tree, r))

    # ── 장애 기록 ─────────────────────────────────────────────────────
    def _record_fault(self, now, ftype, cause, detail):
        row = [now, ftype, cause, detail]
        self._wcsv(self._fault_path, row)
        self.root.after(0, lambda r=row: self._append_ui(self._fault_tree, r, "FAIL"))

    # ── 일별 보고서 ───────────────────────────────────────────────────
    def _generate_daily_report(self, snap=None):
        """CSV 일별 요약 저장."""
        try:
            if snap is None:
                with self._lock:
                    snap = self._stats.snapshot()
            self._wcsv(self._report_path, snap.to_row())
        except Exception:
            pass

    def _generate_excel_report(self, snap=None):
        """Excel(.xlsx) 일별 분석 보고서 생성."""
        try:
            import openpyxl
            from openpyxl.styles import (Alignment, Border, Font,
                                         PatternFill, Side)
        except ImportError:
            return  # openpyxl 미설치 시 스킵

        try:
            if snap is None:
                with self._lock:
                    snap = self._stats.snapshot()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "일별 보고서"

            # ── 스타일 정의 ────────────────────────────────────────
            def fill(hex_color):
                return PatternFill(fgColor=hex_color, fill_type="solid")

            def border():
                s = Side(style="thin", color="CCCCCC")
                return Border(left=s, right=s, top=s, bottom=s)

            C_TITLE  = fill("2E86AB")   # 진파랑
            C_HEAD   = fill("D6EAF8")   # 연파랑
            C_OK     = fill("D5F5E3")   # 연초록
            C_WARN   = fill("FEF9E7")   # 연노랑
            C_CRIT   = fill("FADBD8")   # 연빨강
            C_STRIPE = fill("F8F9FA")   # 연회색

            F_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
            F_HEAD  = Font(name="맑은 고딕", size=10, bold=True, color="1A5276")
            F_BODY  = Font(name="맑은 고딕", size=10)
            F_CRIT  = Font(name="맑은 고딕", size=10, bold=True, color="C0392B")

            CTR = Alignment(horizontal="center", vertical="center")
            LFT = Alignment(horizontal="left",   vertical="center")

            # ── 제목 ───────────────────────────────────────────────
            ws.merge_cells("A1:D1")
            ws["A1"] = "핑감지 테스트기 — 일별 분석 보고서"
            ws["A1"].font      = F_TITLE
            ws["A1"].fill      = C_TITLE
            ws["A1"].alignment = CTR
            ws.row_dimensions[1].height = 32

            # ── 메타 정보 ──────────────────────────────────────────
            ws["A3"] = "보고 일자"
            ws["B3"] = str(snap.date)
            ws["A4"] = "생성 시간"
            ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for r in (3, 4):
                ws.cell(r, 1).font = Font(name="맑은 고딕", size=9, bold=True,
                                          color="555555")
                ws.cell(r, 2).font = Font(name="맑은 고딕", size=9, color="333333")

            # ── 헤더 행 ────────────────────────────────────────────
            ws.row_dimensions[6].height = 22
            for col, txt in enumerate(["항목", "수치", "단위", "평가"], start=1):
                c = ws.cell(6, col, value=txt)
                c.font = F_HEAD; c.fill = C_HEAD
                c.alignment = CTR; c.border = border()

            # ── 데이터 정의 ────────────────────────────────────────
            def _eval(val, warn_thr, crit_thr, lower_is_bad=True):
                """값 평가: 낮을수록 좋은 지표는 lower_is_bad=False."""
                if lower_is_bad:
                    if val == 0:   return "정상", C_OK
                    if val <= warn_thr: return "주의", C_WARN
                    return "경고", C_CRIT
                else:
                    if val <= warn_thr: return "정상", C_OK
                    if val <= crit_thr: return "주의", C_WARN
                    return "경고", C_CRIT

            rows_data = [
                ("총 Ping 수",       snap.total_ping,  "회",  None),
                ("FAIL 횟수",        snap.total_fail,  "회",  _eval(snap.total_fail,  5, 20)),
                ("  └ 설비 Ping 실패", snap.equip_fail, "회",  _eval(snap.equip_fail,  3, 10)),
                ("  └ 서버 Ping 실패", snap.server_fail,"회",  _eval(snap.server_fail, 3, 10)),
                ("최대 응답시간",    snap.max_rt,      "ms",  _eval(snap.max_rt,    200, 500,
                                                                     lower_is_bad=False)),
                ("평균 응답시간",    snap.avg_rt,      "ms",  _eval(snap.avg_rt,     50, 200,
                                                                     lower_is_bad=False)),
                ("NC Agent 종료",    snap.agent_stop,  "회",  _eval(snap.agent_stop,  1,  5)),
                ("동시 장애",        snap.simul_fail,  "회",  _eval(snap.simul_fail,  1,  3)),
            ]

            for row_idx, (label, value, unit, ev) in enumerate(rows_data, start=7):
                ws.row_dimensions[row_idx].height = 20
                is_sub = label.startswith("  └")

                c_lbl = ws.cell(row_idx, 1, value=label)
                c_val = ws.cell(row_idx, 2, value=value)
                c_unt = ws.cell(row_idx, 3, value=unit)

                bg = (ev[1] if ev else
                      C_STRIPE if row_idx % 2 == 0 else
                      PatternFill(fill_type=None))

                for c in (c_lbl, c_val, c_unt):
                    c.font      = F_BODY if not (ev and ev[0] == "경고") else F_CRIT
                    c.fill      = bg
                    c.border    = border()
                    c.alignment = LFT if c is c_lbl else CTR

                if is_sub:
                    c_lbl.font = Font(name="맑은 고딕", size=9, color="666666")

                if ev:
                    c_ev = ws.cell(row_idx, 4, value=ev[0])
                    c_ev.font      = (F_CRIT if ev[0] == "경고" else F_BODY)
                    c_ev.fill      = ev[1]
                    c_ev.alignment = CTR
                    c_ev.border    = border()

            # ── 열 너비 ────────────────────────────────────────────
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 10

            # ── 저장 ───────────────────────────────────────────────
            fname   = f"daily_report_{snap.date.strftime('%Y%m%d')}.xlsx"
            outpath = os.path.join(self._ld, fname)
            wb.save(outpath)

            self.root.after(0, lambda p=outpath: self._safe_status(
                f"Excel 보고서 생성 완료: {p}"))
        except Exception:
            pass

    def _check_day_rollover(self):
        """자정 넘으면 어제 데이터로 보고서 생성 후 통계 초기화."""
        with self._lock:
            if self._stats.date == date.today():
                return
            snap = self._stats.snapshot()
            self._stats.reset()

        # 락 밖에서 파일 IO 수행
        self._generate_daily_report(snap)
        self._generate_excel_report(snap)
        # 응답 누락 분석 집계 초기화
        self._analysis_today.update(
            {"total": 0, "transient": 0, "fault": 0,
             "max_streak": 0, "max_dur": 0, "common": False})
        self.root.after(0, self._refresh_stat_labels)

    # ── 대시보드 통계 갱신 ────────────────────────────────────────────
    def _refresh_stat_labels(self):
        try:
            with self._lock:
                s = self._stats
            for key, val, warn in [
                ("s_ping",   s.total_ping,  False),
                ("s_equip",  s.equip_fail,  s.equip_fail  > 0),
                ("s_server", s.server_fail, s.server_fail > 0),
                ("s_agent",  s.agent_stop,  s.agent_stop  > 0),
                ("s_net",    s.net_error,   s.net_error   > 0),
            ]:
                self._dash[key].config(text=str(val),
                    foreground="#cc0000" if warn else "#006600")
        except tk.TclError:
            pass
        self._update_cause_panel()

    def _update_cause_panel(self):
        """장애 원인 추정 패널을 현재 상태로 갱신."""
        try:
            targets = list(self._targets)
            def _host(t): return str(t[1]) if isinstance(t,(list,tuple)) else t.get("host","")
            def _name(t): return str(t[0]) if isinstance(t,(list,tuple)) else t.get("name","")
            eq_host = _host(targets[0]) if targets           else None
            sv_host = _host(targets[1]) if len(targets) > 1  else None
            eq_name = _name(targets[0]) if targets           else "설비"
            sv_name = _name(targets[1]) if len(targets) > 1  else "서버"

            eq_ok  = self._prev_ok.get(eq_host,  True) if eq_host  else True
            sv_ok  = self._prev_ok.get(sv_host,  True) if sv_host  else True
            ag_up  = self._ncagent_was_up is True
            lk_up  = any(self._prev_if_up.values()) if self._prev_if_up else True

            # ── 상태 표시 ──────────────────────────────────────────
            def _si(key, ok, t_ok, t_fail):
                self._dash[key].config(
                    text=t_ok if ok else t_fail,
                    foreground="#006600" if ok else "#cc0000")

            _si("ce_equip",   eq_ok, f"정상  ({eq_host or '—'})",
                              f"실패  ({eq_host or '—'})")
            _si("ce_server",  sv_ok, f"정상  ({sv_host or '—'})",
                              f"실패  ({sv_host or '—'})")
            _si("ce_agent",   ag_up,
                "실행중" if PSUTIL_OK else "미확인",
                "종료됨" if PSUTIL_OK else "미확인")
            _si("ce_network", lk_up, "Link Up", "Link Down")

            # ── 원인 키 결정 ───────────────────────────────────────
            if not lk_up:
                key = "link_down"
            elif not eq_ok and not sv_ok:
                key = "all_fail"
            elif not eq_ok:
                key = "equip_fail"
            elif not sv_ok:
                key = "server_fail"
            elif not ag_up and PSUTIL_OK:
                key = "agent_only"
            else:
                key = "normal"

            title, cause, action, color = CAUSE_MATRIX[key]

            self._dash["ce_title"].config(
                text=f"[ {title} ]", foreground=color)
            self._dash["ce_cause"].config(text=cause)
            self._dash["ce_action"].config(text=action)

        except (tk.TclError, KeyError):
            pass

    # ── 모니터링 루프들 ───────────────────────────────────────────────
    def _ping_loop(self):
        """
        지수 백오프 + 영구 자동 복구.
        - 30초 이상 정상 실행 후 예외 발생 → 일시적 오류로 간주, streak 리셋
        - 연속 실패 시 2→4→8→...→300초 간격으로 재시도 (영구 중단 없음)
        """
        fail_streak  = 0
        MAX_BACKOFF  = 300  # 최대 5분 대기

        while self._running:
            body_start = time.monotonic()
            try:
                self._ping_loop_body()
                # 정상 종료 (self._running = False) → 루프 자연 종료
            except Exception as e:
                if self._sys_log:
                    self._sys_log.log("_ping_loop", e)

                # 30초 이상 정상 실행 후 실패 → 일시적 오류, streak 리셋
                if time.monotonic() - body_start > 30:
                    fail_streak = 0

                fail_streak += 1
                wait = min(2 * (2 ** min(fail_streak - 1, 7)), MAX_BACKOFF)
                msg  = (f"[경고] Ping 루프 오류 ({fail_streak}회 연속). "
                        f"{wait:.0f}초 후 자동 재시도...")
                self.root.after(0, lambda m=msg: self._safe_status(m))
                self._last_ping_time = time.monotonic()  # watchdog 갱신
                time.sleep(wait)

    def _ping_loop_body(self):
        while self._running:
            now_dt   = datetime.now()
            now      = now_dt.strftime("%Y-%m-%d %H:%M:%S")
            snapshot = list(self._targets)

            # ── Watchdog heartbeat ────────────────────────────────────
            self._last_ping_time = time.monotonic()

            # ── 절전 복귀(시간 점프) 감지 ────────────────────────────
            exp = self._next_cycle_exp
            if exp is not None:
                jump_sec = (now_dt - exp).total_seconds()
                if jump_sec > max(self._interval.get() * 3, 30):
                    # 예상보다 30초+ 늦음 → 절전 복귀 추정
                    self._start_grace_period(45)
            self._next_cycle_exp = (now_dt +
                                    timedelta(seconds=self._interval.get()))

            # ── 이번 사이클 Ping 결과 수집 ────────────────────────────
            prev_snapshot = dict(self._prev_ok)
            results = []
            for i, t in enumerate(snapshot):
                name, host, role = self._t(t)
                if not self._running:
                    break
                ok, rt = self._ping(host)
                results.append((i, name, host, ok, rt))

            # ── 통계 일괄 업데이트 (lock 한 번) ───────────────────────
            with self._lock:
                cycle_fail = sum(1 for _, _, _, ok, _ in results if not ok)
                if cycle_fail == len(results) and len(results) > 0:
                    self._stats.simul_fail += 1   # 전체 동시 FAIL

                for i, name, host, ok, rt in results:
                    self._stats.total_ping += 1
                    if not ok:
                        if i == 0: self._stats.equip_fail += 1
                        else:      self._stats.server_fail += 1
                    elif rt is not None and rt > 0:
                        if rt > self._stats.max_rt:
                            self._stats.max_rt = rt
                        self._stats.sum_rt   += rt
                        self._stats.rt_count += 1

            # ── UI 업데이트 + 장애 기록 + FaultEngine 업데이트 ──────────
            # now_dt는 루프 시작부에서 이미 정의됨
            any_fail = any(not ok for _, _, _, ok, _ in results)

            for i, name, host, ok, rt in results:
                status  = "OK" if ok else "FAIL"
                rt_val  = rt if rt is not None else ""
                rt_disp = f"{rt} ms" if rt is not None else "시간초과"

                # CSV 로그 (core.py CsvLogger 또는 기존 방식)
                if self._csv_log:
                    self._csv_log.write_ping(name, host, ok, rt)
                else:
                    self._wcsv(self._ping_path, [now, host, status, rt_val])

                prev_ok = self._prev_ok.get(host, True)

                # FaultEngine 업데이트 → 상태 변경 이벤트 처리
                eng    = self._engines.get(host) if _CORE_OK else None
                events = eng.update(ok, rt, now_dt) if eng else []
                eng_state = eng.state if eng else (status)

                for ev in events:
                    etype = ev.get("type", "")
                    grace = self._in_grace_period()   # 절전 복귀 유예 기간

                    if etype == "FAULT":
                        if grace:
                            # 유예 기간 중: ping_log에만 기록, fault/스크린샷 생략
                            self.root.after(0, lambda: self._safe_status(
                                f"[유예기간] 절전 복귀 후 네트워크 초기화 중..."))
                        else:
                            self._record_fault(now, "장애 발생",
                                               f"{name} {eng_state}",
                                               f"{host} 연속 {ev.get('fail_streak',0)}회 실패")
                            threading.Thread(target=self._take_screenshot,
                                             daemon=True).start()
                            self.root.after(0, lambda:
                                self._set_led("led_status", "#ff3333"))
                    elif etype == "SUSPECT":
                        if not grace:
                            self._record_fault(now, "장애 의심",
                                               f"{name} 연속 {ev.get('fail_streak',0)}회 실패",
                                               host)
                    elif etype == "RECOVERED":
                        dur_s = ev.get("duration", 0)
                        dur   = (f"{dur_s//60}분 {dur_s%60}초"
                                 if dur_s >= 60 else f"{dur_s}초")
                        rec_row = [now, "Ping 복구", f"{name} 복구",
                                   f"{host}  |  지속시간: {dur}"]
                        self._wcsv(self._fault_path, rec_row)
                        self.root.after(0, lambda r=rec_row:
                            self._append_ui(self._fault_tree, r, "OK"))
                        if not any_fail:
                            self.root.after(0, lambda:
                                self._set_led("led_status", "#ffcc00"))

                # core 없는 경우 기존 방식도 유지
                if not _CORE_OK:
                    if not ok:
                        cause = "설비 Ping FAIL" if i == 0 else "서버 Ping FAIL"
                        self._record_fault(now, "Ping 실패", cause,
                                           f"{host} 응답 없음")
                        if prev_ok:
                            self._fail_start[host] = now_dt
                            threading.Thread(target=self._take_screenshot,
                                             daemon=True).start()
                            self.root.after(0, lambda:
                                self._set_led("led_status", "#ff3333"))
                    elif not prev_ok:
                        start_dt = self._fail_start.pop(host, None)
                        if start_dt:
                            sec = int((now_dt - start_dt).total_seconds())
                            dur = (f"{sec//60}분 {sec%60}초" if sec >= 60
                                   else f"{sec}초")
                            rec_row = [now, "Ping 복구", f"{name} 복구",
                                       f"{host}  |  지속시간: {dur}"]
                            self._wcsv(self._fault_path, rec_row)
                            self.root.after(0, lambda r=rec_row:
                                self._append_ui(self._fault_tree, r, "OK"))
                        if not any_fail:
                            self.root.after(0, lambda:
                                self._set_led("led_status", "#ffcc00"))

                self._prev_ok[host] = ok

                dk = "equip_status" if i == 0 else "server_status"
                dr = "equip_detail" if i == 0 else "server_detail"

                eng_ref = self._engines.get(host) if _CORE_OK else None

                def _upd(s=status, rd=rt_disp, k=dk, kr=dr,
                         n=now, h=host, rv=rt_val, er=eng_ref):
                    try:
                        tag = s if not er else (
                            "OK"   if er.state == FaultEngine.S_NORMAL else
                            "WARN" if er.state in (FaultEngine.S_MISS,
                                                   FaultEngine.S_SUSPECT) else "FAIL")
                        self._append_ui(self._ping_tree, [n, h, s, str(rv)], tag)
                        disp_text = er.state if er else s
                        disp_fg   = er.fg_color if er else (
                            "#006600" if s == "OK" else "#cc0000")
                        disp_bg   = er.bg_color if er else "#FFFFFF"
                        self._dash[k].config(text=disp_text,
                                             foreground=disp_fg,
                                             background=disp_bg)
                        detail = rd
                        if er:
                            detail = (f"{rd}  연속실패:{er.fail_streak}"
                                      if er.fail_streak > 0 else rd)
                        self._dash[kr].config(text=detail)
                    except Exception as exc:
                        if self._sys_log:
                            self._sys_log.log("_upd_dashboard", exc)

                self.root.after(0, _upd)

                # 그래프 데이터 push (첫 두 대상만)
                if i == 0 and self._graph_equip:
                    self._graph_equip.push(rt)
                elif i == 1 and self._graph_server:
                    self._graph_server.push(rt)

            self.root.after(0, self._refresh_stat_labels)
            self.root.after(0, self._refresh_graphs)
            self._analyze_streaks(results, now_dt, prev_snapshot)
            self._check_day_rollover()

            if self._running:
                self.root.after(0, lambda n=now: self._safe_status(
                    f"모니터링 중...   마지막 체크: {n}"))
                time.sleep(self._interval.get())

    def _process_loop(self):
        time.sleep(2)
        while self._running:
            self._check_process()
            for _ in range(5):
                if not self._running:
                    return
                time.sleep(1)

    def _network_loop(self):
        time.sleep(3)
        while self._running:
            self._check_network()
            for _ in range(5):
                if not self._running:
                    return
                time.sleep(1)

    def _event_loop(self):
        time.sleep(15)
        while self._running:
            self._collect_events()
            for _ in range(60):
                if not self._running:
                    return
                time.sleep(1)

    def _ncagent_log_loop(self):
        time.sleep(5)
        while self._running:
            self._check_ncagent_logs()
            for _ in range(30):
                if not self._running:
                    return
                time.sleep(1)

    # ── 스크린샷 (.NET / PowerShell) ─────────────────────────────────
    def _take_screenshot(self):
        """장애 발생 시 PowerShell(.NET)로 전체 화면 캡처."""
        try:
            sdir = os.path.join(self._ld, "screenshots")
            os.makedirs(sdir, exist_ok=True)
            fname = f"Screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            fpath = os.path.join(sdir, fname).replace("'", "''")

            ps = (
                "Add-Type -AssemblyName System.Windows.Forms,System.Drawing; "
                "$s=[System.Windows.Forms.Screen]::PrimaryScreen.Bounds; "
                "$b=New-Object System.Drawing.Bitmap($s.Width,$s.Height); "
                "$g=[System.Drawing.Graphics]::FromImage($b); "
                "$g.CopyFromScreen($s.Location,"
                "[System.Drawing.Point]::Empty,$s.Size); "
                f"$b.Save('{fpath}'); "
                "$g.Dispose();$b.Dispose()"
            )
            subprocess.run(
                ["powershell", "-NonInteractive", "-NoProfile", "-Command", ps],
                capture_output=True, timeout=12,
                creationflags=subprocess.CREATE_NO_WINDOW)
        except Exception:
            pass

    # ── LED 인디케이터 ────────────────────────────────────────────────
    def _set_led(self, key, color):
        """대시보드 LED 색상 변경."""
        try:
            self._dash[key].itemconfig("led", fill=color)
        except (tk.TclError, KeyError):
            pass

    # ── 로그 자동 정리 ────────────────────────────────────────────────
    # 아카이브 대상 CSV 파일 (헤더 상수 이름)
    _ARCHIVE_CSV = [
        ("ping_log.csv",    "H_PING"),
        ("process_log.csv", "H_PROC"),
        ("network_log.csv", "H_NET"),
        ("event_log.csv",   "H_EVT"),
        ("fail_log.csv",    "H_FAULT"),
        ("fault_log.csv",   "H_FAULT"),   # core.py 새 형식
        ("agent_log.csv",   "H_PROC"),    # core.py 새 형식
    ]

    def _archive_old_logs(self):
        """
        보관 기간 초과 로그를 archive 폴더로 이동 (Streaming + _log_lock).
        메모리에 전체 파일을 올리지 않고 임시 파일 방식으로 처리.
        """
        days    = self._retention_days.get()
        cutoff  = datetime.now() - timedelta(days=days)
        arc_dir = os.path.join(self._ld, "archive")
        total   = 0

        for fname, hdr_name in self._ARCHIVE_CSV:
            total += self._archive_csv_stream(fname, hdr_name, cutoff, arc_dir)

        total += self._archive_error_log(cutoff, arc_dir)
        total += self._archive_screenshots(cutoff)
        return total

    def _archive_csv_stream(self, fname, hdr_name, cutoff, arc_dir):
        """단일 CSV를 스트리밍 방식으로 아카이브 (파일 전체를 메모리에 올리지 않음)."""
        src = os.path.join(self._ld, fname)
        if not os.path.exists(src):
            return 0

        tmp          = src + ".archtmp"
        arc_handles  = {}   # month -> (file_obj, csv_writer)
        archived     = 0
        default_hdr  = globals().get(hdr_name, [])

        try:
            os.makedirs(arc_dir, exist_ok=True)
            with self._log_lock:  # Ping 쓰기 스레드 차단 후 atomic 처리
                fin  = open(src, "r", encoding="utf-8-sig", newline="")
                ftmp = open(tmp, "w", encoding="utf-8-sig", newline="")
                try:
                    reader   = csv.reader(fin)
                    writer   = csv.writer(ftmp)
                    file_hdr = next(reader, None) or default_hdr
                    writer.writerow(file_hdr)

                    for row in reader:
                        if not row:
                            continue
                        try:
                            dt = datetime.strptime(row[0][:19], "%Y-%m-%d %H:%M:%S")
                        except (ValueError, IndexError):
                            writer.writerow(row)  # 날짜 없는 행은 유지
                            continue

                        if dt < cutoff:
                            month = row[0][:7].replace("-", "")
                            if month not in arc_handles:
                                arc_p  = os.path.join(arc_dir,
                                    f"{fname.replace('.csv','')}"
                                    f"_{month}.csv")
                                exists = os.path.exists(arc_p)
                                fh     = open(arc_p, "a",
                                              encoding="utf-8-sig", newline="")
                                aw     = csv.writer(fh)
                                if not exists:
                                    aw.writerow(file_hdr)
                                arc_handles[month] = (fh, aw)
                            arc_handles[month][1].writerow(row)
                            archived += 1
                        else:
                            writer.writerow(row)
                finally:
                    fin.close()
                    ftmp.close()
                    for fh, _ in arc_handles.values():
                        try:
                            fh.close()
                        except Exception:
                            pass

                if archived > 0:
                    os.replace(tmp, src)
                else:
                    os.remove(tmp)

        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_archive_csv_stream", e, fname)
            for cleanup in [tmp]:
                try:
                    if os.path.exists(cleanup):
                        os.remove(cleanup)
                except Exception:
                    pass

        return archived

    def _archive_error_log(self, cutoff, arc_dir):
        """system_error.log에서 오래된 블록을 archive로 이동."""
        src = os.path.join(self._ld, "system_error.log")
        if not os.path.exists(src):
            return 0

        tmp      = src + ".archtmp"
        archived = 0

        try:
            os.makedirs(arc_dir, exist_ok=True)
            with self._log_lock:
                old_lines, new_lines = [], []
                cur_block  = []
                cur_is_old = False

                with open(src, "r", encoding="utf-8", errors="replace") as f:
                    for line in f:
                        if line.startswith("[20") and len(line) > 20:
                            if cur_block:
                                (old_lines if cur_is_old else new_lines).extend(cur_block)
                            try:
                                ts = datetime.strptime(line[1:20], "%Y-%m-%d %H:%M:%S")
                                cur_is_old = ts < cutoff
                            except (ValueError, IndexError):
                                cur_is_old = False
                            cur_block = [line]
                        else:
                            cur_block.append(line)
                    if cur_block:
                        (old_lines if cur_is_old else new_lines).extend(cur_block)

                if old_lines:
                    month = old_lines[0][1:8].replace("-", "")
                    arc_p = os.path.join(arc_dir, f"system_error_{month}.log")
                    with open(arc_p, "a", encoding="utf-8") as f:
                        f.writelines(old_lines)
                    archived = sum(1 for l in old_lines if l.startswith("[20"))

                with open(tmp, "w", encoding="utf-8") as f:
                    f.writelines(new_lines)
                os.replace(tmp, src)

        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_archive_error_log", e)
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass

        return archived

    def _archive_screenshots(self, cutoff):
        """보관 기간 초과 스크린샷 PNG 자동 삭제."""
        sdir    = os.path.join(self._ld, "screenshots")
        deleted = 0
        if not os.path.isdir(sdir):
            return 0
        try:
            for fname in os.listdir(sdir):
                if not fname.lower().endswith(".png"):
                    continue
                fpath = os.path.join(sdir, fname)
                try:
                    mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                    if mtime < cutoff:
                        os.remove(fpath)
                        deleted += 1
                except Exception:
                    pass
        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_archive_screenshots", e)
        return deleted

    def _run_archive_now(self):
        """'지금 정리' 버튼 → 별도 스레드에서 실행."""
        self._archive_btn.config(state=tk.DISABLED)
        self._archive_status.config(text="정리 중...")

        def _do():
            try:
                n = self._archive_old_logs()
                msg = (f"{n}건 아카이브 완료" if n > 0
                       else "이동할 항목 없음")
            except Exception as e:
                msg = f"오류: {e}"
            self.root.after(0, lambda m=msg: (
                self._archive_status.config(text=m),
                self._archive_btn.config(state=tk.NORMAL),
            ))

        threading.Thread(target=_do, daemon=True).start()

    def _archive_loop(self):
        """시작 후 1분 대기, 이후 24시간마다 자동 정리."""
        time.sleep(60)
        while self._running:
            self._archive_old_logs()
            self.root.after(0, lambda: self._archive_status.config(
                text=f"마지막 자동 정리: {datetime.now().strftime('%m-%d %H:%M')}"))
            for _ in range(1440):       # 24시간 = 1440분
                if not self._running:
                    return
                time.sleep(60)

    # ── 응답 누락 분석 ────────────────────────────────────────────────
    def _analyze_streaks(self, results, now_dt, prev_snapshot):
        """연속 실패 그룹을 추적하고, 복구 시 자동 분류."""
        if not results:
            return

        targets    = list(self._targets)
        result_map = {host: ok for _, _, host, ok, _ in results}

        # 이번 사이클에 전체 대상이 동시 실패했는지 확인
        all_fail = (len(results) > 1 and
                    all(not ok for _, _, _, ok, _ in results))

        for i, t in enumerate(targets):
            # tuple / dict / list 모두 지원
            if isinstance(t, (list, tuple)):
                name, host = str(t[0]), str(t[1]) if len(t) > 1 else ""
            else:
                name, host = t.get("name",""), t.get("host","")

            if host not in result_map:
                continue
            ok      = result_map[host]
            was_ok  = prev_snapshot.get(host, True)

            if not ok:
                if was_ok:                      # OK → FAIL : 연속 실패 시작
                    self._streak_count[host]  = 1
                    self._streak_start[host]  = now_dt
                    self._streak_simul[host]  = all_fail
                    # 서버 Ping 상태 기록 (설비 자신이 아닌 다른 첫 번째 대상)
                    def _get_host(tt):
                        return str(tt[1]) if isinstance(tt,(list,tuple)) else tt.get("host","")
                    other = next((_get_host(targets[j]) for j in range(len(targets)) if j != i), None)
                    self._streak_svr_ok[host] = result_map.get(other, True) if other else True
                else:                           # FAIL 지속
                    self._streak_count[host] = self._streak_count.get(host, 0) + 1
                    if all_fail:
                        self._streak_simul[host] = True
            else:
                if not was_ok:                  # FAIL → OK : 복구 → 분류
                    streak   = self._streak_count.get(host, 1)
                    start    = self._streak_start.get(host, now_dt)
                    svr_ok   = self._streak_svr_ok.get(host, True)
                    simul    = self._streak_simul.get(host, False)
                    dur_sec  = max(1, int((now_dt - start).total_seconds()))
                    self._record_analysis(name, host, streak, dur_sec,
                                          svr_ok, simul, start, now_dt)
                    for d in (self._streak_count, self._streak_start,
                              self._streak_svr_ok, self._streak_simul):
                        d.pop(host, None)

    def _record_analysis(self, name, host, streak, dur_sec,
                          svr_ok, simul, fail_start, fail_end):
        """분석 결과 분류 후 테이블·요약 업데이트."""
        ts_str   = fail_end.strftime("%Y-%m-%d %H:%M:%S")
        dur_str  = (f"{dur_sec//60}분 {dur_sec%60}초"
                    if dur_sec >= 60 else f"{dur_sec}초")
        svr_text = "정상" if svr_ok else "실패"
        eq_text  = "실패→복구"

        # ── 판정 ──────────────────────────────────────────────────
        if simul or not svr_ok:
            verdict = "공통 구간 장애 가능성"
            cause   = ("서버·설비 Ping 동시 실패\n"
                       "→ PC/랜어댑터 또는 상위 네트워크 문제 가능성")
            action  = "공유기·스위치 재시작, PC 네트워크 설정 확인"
            tag     = "COMMON"
        elif streak <= 2:
            verdict = (f"{streak}회 실패 후 즉시 복구 — 일시적 응답 누락")
            cause   = "패킷 순간 손실 (네트워크 혼잡 또는 일시적 지연)"
            action  = "빈도 증가 시 점검 권장, 현재는 지속 모니터링"
            tag     = "TRANSIENT"
        else:
            verdict = (f"{streak}회 연속 실패 — 실제 네트워크 장애 가능성")
            cause   = ("서버 Ping 정상, 설비 Ping 실패\n"
                       "→ 설비/스위치/현장망 구간 문제 가능성")
            action  = "설비 전원·케이블·스위치 포트 점검"
            tag     = "FAULT"

        # ── 오늘 집계 업데이트 ─────────────────────────────────────
        self._analysis_today["total"] += 1
        if tag == "TRANSIENT":
            self._analysis_today["transient"] += 1
        else:
            self._analysis_today["fault"] += 1
        self._analysis_today["max_streak"] = max(
            self._analysis_today["max_streak"], streak)
        self._analysis_today["max_dur"] = max(
            self._analysis_today["max_dur"], dur_sec)
        if simul:
            self._analysis_today["common"] = True

        row = [ts_str, name, f"{streak}회", dur_str,
               svr_text, eq_text, verdict, cause, action]

        self.root.after(0, lambda r=row, t=tag: (
            self._append_ui(self._analysis_tree, r, t),
            self._update_analysis_summary_ui(),
        ))

    def _update_analysis_summary_ui(self):
        """분석 요약 박스 레이블 업데이트."""
        try:
            a = self._analysis_today
            self._dash["as_total"].config(text=str(a["total"]))
            self._dash["as_transient"].config(
                text=str(a["transient"]),
                foreground="#cc8800" if a["transient"] > 0 else "#006600")
            self._dash["as_fault"].config(
                text=str(a["fault"]),
                foreground="#cc0000" if a["fault"] > 0 else "#006600")
            ms = a["max_streak"]
            self._dash["as_max_streak"].config(
                text=f"{ms}회" if ms > 0 else "—",
                foreground="#cc0000" if ms >= 3 else "#333333")
            md = a["max_dur"]
            self._dash["as_max_dur"].config(
                text=(f"{md//60}분 {md%60}초" if md >= 60 else f"{md}초")
                if md > 0 else "—")
            common = a["common"]
            self._dash["as_common"].config(
                text="발생" if common else "없음",
                foreground="#cc0000" if common else "#006600")
        except tk.TclError:
            pass

    def _refresh_graphs(self):
        """그래프 갱신 — 메인 스레드에서 호출."""
        try:
            if self._graph_equip:  self._graph_equip.refresh()
            if self._graph_server: self._graph_server.refresh()
        except tk.TclError:
            pass

    # ════════════════════════════════════════════════════════════════
    # 24시간 안정성 메서드
    # ════════════════════════════════════════════════════════════════

    # ── NC Agent 경로 직접 선택 ──────────────────────────────────
    def _browse_ncagent_dir(self) -> None:
        """NC Agent 설치 경로를 사용자가 직접 선택."""
        cur = self._ncagent_dir or os.path.expanduser("~")
        if not os.path.isdir(cur):
            cur = os.path.expanduser("~")
        d = filedialog.askdirectory(
            title="NC Agent 설치 폴더 선택 (Log 폴더 또는 설치 폴더)",
            initialdir=cur)
        if not d:
            return
        # Log 하위 폴더가 있으면 그쪽 사용
        log_sub = os.path.join(d, "Log")
        self._ncagent_dir = log_sub if os.path.isdir(log_sub) else d
        self._ncagent_log_mtimes.clear()  # 재스캔
        self._ncagent_path_lbl.config(
            text=self._ncagent_dir, foreground="#333333")
        # config에 저장 (재시작해도 유지)
        self._save_config()

    # ── psutil 설치 버튼 위젯 ─────────────────────────────────────
    def _build_psutil_install_bar(self, parent) -> None:
        """psutil 미설치 경고 + 설치 버튼."""
        bar = ttk.Frame(parent)
        bar.pack(fill=tk.X, padx=12, pady=6)
        ttk.Label(bar,
                  text="psutil 미설치: NC Agent / 네트워크 어댑터 기능 비활성화",
                  foreground="#cc6600").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(bar, text="지금 설치",
                   command=self._install_psutil, width=12).pack(side=tk.LEFT)

    def _install_psutil(self) -> None:
        """psutil을 pip으로 자동 설치."""
        self.root.config(cursor="wait")
        self.root.update()
        try:
            import subprocess
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", "psutil"],
                capture_output=True, text=True, timeout=60,
                creationflags=subprocess.CREATE_NO_WINDOW)

            if result.returncode == 0:
                messagebox.showinfo(
                    "psutil 설치 완료",
                    "psutil 설치가 완료되었습니다.\n\n"
                    "프로그램을 종료 후 다시 실행하면\n"
                    "NC Agent 감시 / 네트워크 어댑터 기능이 활성화됩니다.")
            else:
                err = (result.stderr or result.stdout or "")[-400:]
                messagebox.showerror(
                    "설치 실패",
                    f"psutil 설치 실패:\n{err}\n\n"
                    "인터넷 연결을 확인하거나\n"
                    "명령 프롬프트에서 직접 실행해보세요:\n"
                    "pip install psutil")
        except Exception as e:
            messagebox.showerror("오류", str(e))
        finally:
            self.root.config(cursor="")

    # ── 절전 복귀 Grace Period ────────────────────────────────────
    def _start_grace_period(self, duration: int = 45) -> None:
        """절전 복귀 후 N초 동안 FAULT 판정·스크린샷·알람 금지."""
        self._grace_until = datetime.now() + timedelta(seconds=duration)
        self.root.after(0, lambda d=duration: self._safe_status(
            f"[절전 복귀 감지] {d}초 유예 기간 적용 — FAULT 판정 보류"))

    def _in_grace_period(self) -> bool:
        return (self._grace_until is not None
                and datetime.now() < self._grace_until)

    # ── 비정상 종료 감지 ──────────────────────────────────────────
    def _mark_running(self) -> None:
        """프로그램 실행 중 마킹 파일 생성."""
        try:
            with open(_FLAG_FILE, "w", encoding="utf-8") as f:
                f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        except Exception:
            pass

    def _clear_running(self) -> None:
        """정상 종료 시 마킹 파일 삭제."""
        try:
            if os.path.exists(_FLAG_FILE):
                os.remove(_FLAG_FILE)
        except Exception:
            pass

    def _check_previous_crash(self) -> None:
        """이전 비정상 종료 여부 확인 → UI에 알림."""
        if not os.path.exists(_FLAG_FILE):
            return
        try:
            with open(_FLAG_FILE, encoding="utf-8") as f:
                prev = f.read().strip()
        except Exception:
            prev = "알 수 없음"
        # 마킹 파일 삭제 (다음 실행에서 중복 알림 방지)
        self._clear_running()
        self.root.after(800, lambda: messagebox.showwarning(
            "이전 세션 비정상 종료",
            f"이전 실행이 정상 종료되지 않았습니다.\n"
            f"마지막 실행 시각: {prev}\n\n"
            f"로그 파일을 확인하세요: {self._ld}"))

    # ── 1분마다 config 자동 저장 ──────────────────────────────────
    def _periodic_config_save(self) -> None:
        """1분마다 config.json 자동 저장 (강제 종료 대비)."""
        try:
            self._save_config()
        except Exception:
            pass
        try:
            self.root.after(60_000, self._periodic_config_save)
        except tk.TclError:
            pass

    # ── 날짜 변경 감시 (모니터링 중지 중에도 작동) ────────────────
    def _watch_day(self) -> None:
        """1분마다 날짜 변경 확인 → 자정 롤오버 처리."""
        try:
            if self._stats.date != date.today():
                self._check_day_rollover()
        except Exception:
            pass
        try:
            self.root.after(60_000, self._watch_day)
        except tk.TclError:
            pass

    # ── Watchdog: Ping 루프 무응답 감지 ──────────────────────────
    def _watchdog_check(self) -> None:
        """15초마다 마지막 Ping 시각 확인 → 30초 무응답 시 경고."""
        try:
            if self._running:
                elapsed = time.monotonic() - self._last_ping_time
                if elapsed > 30:
                    self._safe_status(
                        f"[Watchdog] Ping 루프 {elapsed:.0f}초 무응답 — 자동 복구 대기 중")
            self.root.after(15_000, self._watchdog_check)
        except tk.TclError:
            pass

    def _safe_status(self, text):
        try:
            self._status_bar.config(text=text)
        except tk.TclError:
            pass

    # ── 시작 / 중지 / 종료 ────────────────────────────────────────────
    def _start(self):
        if not self._targets:
            messagebox.showwarning("알림", "대상 IP를 먼저 추가하세요.")
            return
        if not self._log_dir.get().strip():
            messagebox.showwarning("알림", "로그 경로를 입력하세요.")
            return

        # 설정 검증
        if _CORE_OK:
            cfg = {"targets": [{"name": self._t(t)[0], "host": self._t(t)[1]}
                                for t in self._targets],
                   "interval": self._interval.get()}
            ok, errs = validate_config(cfg)
            if not ok:
                messagebox.showerror("설정 오류", "\n".join(errs))
                return
            warn = validate_interval_warn(self._interval.get())
            if warn:
                if not messagebox.askyesno("검사 주기 경고", warn + "\n\n계속 시작하시겠습니까?"):
                    return

        try:
            self._ensure_logs()
        except Exception as e:
            if self._sys_log:
                self._sys_log.log("_start._ensure_logs", e)
            messagebox.showerror("오류", f"로그 폴더 생성 실패:\n{e}")
            return

        # core 모듈 경로 업데이트
        if _CORE_OK:
            self._sys_log.update_path(self._ld)
            self._csv_log.update_dir(self._ld)
            rpt_dir = os.path.join(_SCRIPT_DIR, "reports")
            os.makedirs(rpt_dir, exist_ok=True)
            self._reporter.update_dir(rpt_dir)

        # FaultEngine 생성 (대상별)
        if _CORE_OK:
            self._engines.clear()
            roles = getattr(self, "_target_roles", {})
            for i, t in enumerate(self._targets):
                name = str(t[0]) if isinstance(t, (list, tuple)) else t.get("name","")
                host = str(t[1]) if isinstance(t, (list, tuple)) else t.get("host","")
                role = roles.get(host, "equipment" if i == 0 else "server")
                if host:
                    self._engines[host] = FaultEngine(
                        name, host, role, self._fault_policy, self._csv_log)

        self._running = True
        self._threads.clear()
        self._start_btn.config(state=tk.DISABLED)
        self._stop_btn.config(state=tk.NORMAL)
        self._add_btn.config(state=tk.DISABLED)
        self._del_btn.config(state=tk.DISABLED)
        self._edt_btn.config(state=tk.DISABLED)
        self._log_entry.config(state=tk.DISABLED)
        self._browse_btn.config(state=tk.DISABLED)

        self._archive_btn.config(state=tk.DISABLED)

        for fn in [self._ping_loop, self._process_loop, self._network_loop,
                   self._event_loop, self._ncagent_log_loop, self._archive_loop]:
            t = threading.Thread(target=fn, daemon=True)
            self._threads.append(t)
            t.start()

        self._set_led("led_power",  "#00dd44")  # 녹색
        self._set_led("led_status", "#ffcc00")  # 노란색

        # 비정상 종료 감지 마킹
        self._mark_running()
        # 1분마다 config 자동 저장
        self.root.after(60_000, self._periodic_config_save)
        # Watchdog 시작
        self.root.after(15_000, self._watchdog_check)
        # next_cycle 초기화 (sleep 감지 오탐 방지)
        self._next_cycle_exp = None

        self._status_bar.config(text=f"모니터링 시작  —  로그: {self._ld}")

    def _stop(self):
        self._running = False
        self._generate_daily_report()
        self._generate_excel_report()
        self._set_led("led_power",  "#888888")  # 회색
        self._set_led("led_status", "#888888")
        self._reset_buttons()
        self._status_bar.config(text="모니터링 중지됨")

    def _reset_buttons(self):
        self._start_btn.config(state=tk.NORMAL)
        self._stop_btn.config(state=tk.DISABLED)
        self._add_btn.config(state=tk.NORMAL)
        self._del_btn.config(state=tk.NORMAL)
        self._edt_btn.config(state=tk.NORMAL)
        self._log_entry.config(state=tk.NORMAL)
        self._browse_btn.config(state=tk.NORMAL)
        self._archive_btn.config(state=tk.NORMAL)

    def _on_close(self):
        # 모니터링 중일 때만 확인 다이얼로그 표시
        if self._running:
            dlg = tk.Toplevel(self.root)
            dlg.title("종료 확인")
            dlg.resizable(False, False)
            dlg.grab_set()
            dlg.transient(self.root)

            ttk.Label(dlg,
                      text="모니터링이 실행 중입니다.\n정말 종료하시겠습니까?",
                      font=("", 10), justify=tk.CENTER,
                      padding=(20, 16)).pack()

            ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)

            btn_row = ttk.Frame(dlg, padding=(0, 10))
            btn_row.pack()

            confirmed = [False]

            def _yes():
                confirmed[0] = True
                dlg.destroy()

            ttk.Button(btn_row, text="종료", command=_yes, width=12).pack(side=tk.LEFT, padx=8)
            if self._tray and self._tray.available:
                def _tray_min():
                    dlg.destroy()
                    self.root.withdraw()
                    self._tray.start("핑감지 테스트기")
                ttk.Button(btn_row, text="트레이 최소화",
                           command=_tray_min, width=14).pack(side=tk.LEFT, padx=8)
            ttk.Button(btn_row, text="취소", command=dlg.destroy, width=12).pack(side=tk.LEFT, padx=8)

            dlg.bind("<Return>", lambda e: _yes())
            dlg.bind("<Escape>", lambda e: dlg.destroy())

            # 창 중앙 배치
            dlg.update_idletasks()
            px = self.root.winfo_x() + (self.root.winfo_width()  - dlg.winfo_width())  // 2
            py = self.root.winfo_y() + (self.root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f"+{px}+{py}")

            self.root.wait_window(dlg)

            if not confirmed[0]:
                return   # 취소 → 종료하지 않음

        self._running = False
        if _CORE_OK and self._reporter and self._engines:
            try:
                self._reporter.generate(list(self._engines.values()))
            except Exception as e:
                if self._sys_log:
                    self._sys_log.log("_on_close.reporter", e)
        self._generate_daily_report()
        self._generate_excel_report()
        self._save_config()
        self._clear_running()   # 정상 종료 마킹 파일 삭제
        if self._tray:
            self._tray.stop()
        self.root.quit()
        self.root.destroy()

    # ── 트레이 아이콘 콜백 ───────────────────────────────────────────
    def _tray_show(self) -> None:
        self.root.after(0, self.root.deiconify)
        self.root.after(0, self.root.lift)

    def _tray_quit(self) -> None:
        self.root.after(0, self._on_close)

    # ── 시작 프로그램 등록 UI (설정 스트립에서 호출) ─────────────────
    def _toggle_startup(self) -> None:
        if not _CORE_OK:
            messagebox.showinfo("알림", "startup.py가 없어 이 기능을 사용할 수 없습니다.")
            return
        enabled = is_startup_enabled()
        ok, msg = set_startup(not enabled)
        if ok:
            self._startup_btn.config(
                text="자동 실행 해제" if not enabled else "시작 시 자동 실행")
        messagebox.showinfo("자동 실행 설정", msg)


# ── 아이콘 생성 ───────────────────────────────────────────────────────
_ICON_PATH = os.path.join(_BASE_DIR, "app_icon.ico")


def _make_icon_png():
    """64×64 귀여운 핑 아이콘 PNG 바이트 생성 (순수 Python)."""
    W = H = 64
    cx = cy = 32.0

    rows = []
    for y in range(H):
        row = bytearray([0])          # PNG filter: None
        for x in range(W):
            dx, dy = x - cx, y - cy
            dist   = math.sqrt(dx * dx + dy * dy)
            R      = 29.0

            if dist > R + 1.5:
                row += b'\x00\x00\x00\x00'
                continue

            # ── 배경 : 하늘색 → 파란색 그라디언트 ──────────────────
            t  = min(dist / R, 1.0)
            rc = int(100 - t * 25)
            gc = int(180 - t * 40)
            bc = int(240 - t * 40)

            # 테두리 부드럽게
            alpha = min(255, max(0, int(255 * (R + 1.5 - dist) / 1.5)))

            # ── 흰색 핑 아크 (상단 반원) ───────────────────────────
            angle    = math.degrees(math.atan2(-dy, dx))  # 90°=위, 0°=오른쪽
            on_white = dist < 4.5   # 중심 점

            if not on_white:
                for arc_r, arc_w in [(9.5, 2.2), (17.0, 2.2), (24.5, 2.2)]:
                    if abs(dist - arc_r) < arc_w and 25 <= angle <= 155:
                        on_white = True
                        break

            if on_white:
                row += bytes([255, 255, 255, alpha])
            else:
                row += bytes([rc, gc, bc, alpha])
        rows.append(bytes(row))

    raw  = b''.join(rows)
    comp = zlib.compress(raw, 6)

    def chunk(name, data):
        body = name + data
        return (struct.pack('>I', len(data)) + body
                + struct.pack('>I', zlib.crc32(body) & 0xFFFFFFFF))

    return (b'\x89PNG\r\n\x1a\n'
            + chunk(b'IHDR', struct.pack('>II', W, H) + bytes([8, 6, 0, 0, 0]))
            + chunk(b'IDAT', comp)
            + chunk(b'IEND', b''))


def _make_icon_ico(png_data):
    """PNG 데이터를 단일 ICO 파일로 래핑 (Windows PNG-in-ICO 지원)."""
    header = struct.pack('<HHH', 0, 1, 1)            # 예약, type=1(ICO), 이미지수=1
    entry  = struct.pack('<BBBBHHII',
                         64, 64,           # width, height
                         0,                # 색상 수 (true-color = 0)
                         0,                # 예약
                         1,                # planes
                         32,               # bpp
                         len(png_data),
                         6 + 16)           # 데이터 오프셋
    return header + entry + png_data


def _setup_icon(root):
    """아이콘 생성 → .ico 저장 → 창 아이콘 적용."""
    try:
        png = _make_icon_png()
        ico = _make_icon_ico(png)

        with open(_ICON_PATH, 'wb') as f:
            f.write(ico)
        root.iconbitmap(_ICON_PATH)
    except Exception:
        # .ico 실패 시 PhotoImage로 대체 시도
        try:
            img = tk.PhotoImage(data=base64.b64encode(png).decode())
            root.iconphoto(True, img)
            root._app_icon = img          # GC 방지
        except Exception:
            pass


if __name__ == "__main__":
    root = tk.Tk()
    _setup_icon(root)
    PingMonitorApp(root)
    root.mainloop()
    try:
        root.destroy()
    except Exception:
        pass
