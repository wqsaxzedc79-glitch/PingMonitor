"""
create_manual.py
핑감지 테스트기 v2 사용자 매뉴얼 생성 스크립트
실행: python create_manual.py
"""

import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
except ImportError:
    print("[오류] openpyxl 필요: pip install openpyxl")
    sys.exit(1)

# ── 색상 팔레트 ───────────────────────────────────────────────────────
C = {
    "navy":    "1F3864",
    "blue":    "2E75B6",
    "blue_lt": "D6E4F0",
    "blue_hl": "BDD7EE",
    "green":   "375623",
    "green_l": "E2EFDA",
    "green_h": "00B050",
    "yellow":  "7F6000",
    "yell_l":  "FFEB9C",
    "yell_h":  "FFC000",
    "red":     "9C0006",
    "red_l":   "FFC7CE",
    "red_h":   "FF0000",
    "purple":  "7030A0",
    "purp_l":  "EAD1F7",
    "gray":    "D9D9D9",
    "gray_d":  "595959",
    "white":   "FFFFFF",
    "black":   "000000",
    "orange":  "843C0C",
    "oran_l":  "FCE4D6",
}

TODAY = datetime.now().strftime("%Y년 %m월 %d일")
OUT   = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "핑감지_테스트기_v2_사용자_매뉴얼.xlsx")

# ── 스타일 헬퍼 ───────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fgColor=hex_color, fill_type="solid")

def font(color="000000", size=10, bold=False, italic=False, name="맑은 고딕"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color="BFBFBF")
    m = Side(style="medium", color="595959")
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "outer":
        return Border(left=m, right=m, top=m, bottom=m)
    return Border()

def thick_border():
    m = Side(style="medium", color="1F3864")
    return Border(left=m, right=m, top=m, bottom=m)

def nav_link(ws, row, col, text, target_sheet, target_cell="A1"):
    c = ws.cell(row=row, column=col, value=text)
    c.hyperlink = f"#'{target_sheet}'!{target_cell}"
    c.font      = Font(name="맑은 고딕", size=9, color="2E75B6",
                       underline="single")
    c.alignment = align("center")
    return c

def set_title(ws, row, text, merge_to_col=16, color=C["navy"], fsize=16):
    ws.row_dimensions[row].height = 36
    c = ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=merge_to_col)
    c.font      = font(C["white"], fsize, bold=True)
    c.fill      = fill(color)
    c.alignment = align("center")

def set_section(ws, row, text, merge_to=16, color=C["blue"]):
    ws.row_dimensions[row].height = 24
    c = ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=merge_to)
    c.font      = font(C["white"], 11, bold=True)
    c.fill      = fill(color)
    c.alignment = align("left")
    return c

def table_header(ws, row, headers, col_start=1, bg=C["navy"]):
    ws.row_dimensions[row].height = 22
    for i, h in enumerate(headers, col_start):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = font(C["white"], 10, bold=True)
        c.fill      = fill(bg)
        c.alignment = align("center")
        c.border    = thin_border()

def table_row(ws, row, values, col_start=1, alt=False, colors=None):
    bg = C["blue_lt"] if alt else C["white"]
    ws.row_dimensions[row].height = 18
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        cell_bg = colors[i - col_start] if colors and i - col_start < len(colors) else bg
        c.fill      = fill(cell_bg)
        c.font      = font(size=9)
        c.alignment = align("left", wrap=True)
        c.border    = thin_border()

def placeholder(ws, r1, c1, r2, c2, text="[ 화면 캡처 삽입 위치 ]"):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=text)
    c.font      = Font(name="맑은 고딕", size=12, color=C["gray_d"], italic=True)
    c.fill      = fill("F2F2F2")
    c.alignment = align("center")
    c.border    = thick_border()
    for r in range(r1, r2 + 1):
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 20)

def add_nav_bar(ws, current, sheets, row=2):
    """상단 네비게이션 바."""
    ws.row_dimensions[row].height = 18
    labels = [s["short"] for s in sheets]
    for i, sh in enumerate(sheets, 1):
        c = ws.cell(row=row, column=i, value=sh["short"])
        if sh["short"] == current:
            c.fill = fill(C["navy"])
            c.font = Font(name="맑은 고딕", size=8, bold=True, color=C["white"])
        else:
            c.hyperlink = f"#'{sh['name']}'!A1"
            c.fill = fill(C["blue_lt"])
            c.font = Font(name="맑은 고딕", size=8, color=C["blue"],
                          underline="single")
        c.alignment = align("center")
        c.border    = thin_border()

def a4_landscape(ws):
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.paperSize     = 9   # A4
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.page_margins             = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3)
    ws.print_title_rows         = "1:3"


# ── 시트 목록 ─────────────────────────────────────────────────────────
SHEETS = [
    {"name": "표지",             "short": "① 표지"},
    {"name": "프로그램 개요",    "short": "② 개요"},
    {"name": "화면 구성 설명",   "short": "③ 화면"},
    {"name": "기능 설명",        "short": "④ 기능"},
    {"name": "장애 분석 기능",   "short": "⑤ 장애"},
    {"name": "운영 매뉴얼",      "short": "⑥ 운영"},
    {"name": "장애 유형별 조치", "short": "⑦ 조치"},
    {"name": "FAQ",              "short": "⑧ FAQ"},
    {"name": "유지보수 이력",    "short": "⑨ 이력"},
]


# ═══════════════════════════════════════════════════════════════════════
# 시트1: 표지
# ═══════════════════════════════════════════════════════════════════════
def build_cover(ws):
    a4_landscape(ws)
    ws.column_dimensions["A"].width = 4

    for i in range(1, 17):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["P"].width = 3

    add_nav_bar(ws, "① 표지", SHEETS, row=1)

    # 상단 장식
    ws.merge_cells("A3:P4")
    c = ws.cell(row=3, column=1)
    c.fill = fill(C["navy"])
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 8

    # 제목 블록
    ws.merge_cells("B6:O6")
    ws.row_dimensions[6].height = 50
    c = ws.cell(row=6, column=2,
                value="핑감지 테스트기 v2")
    c.font      = Font(name="맑은 고딕", size=32, bold=True, color=C["navy"])
    c.alignment = align("center")

    ws.merge_cells("B7:O7")
    ws.row_dimensions[7].height = 28
    c = ws.cell(row=7, column=2,
                value="사용자 매뉴얼 (User Manual)")
    c.font      = Font(name="맑은 고딕", size=18, color=C["blue"])
    c.alignment = align("center")

    ws.merge_cells("B8:O8")
    ws.row_dimensions[8].height = 28
    c = ws.cell(row=8, column=2,
                value="NC Agent 네트워크 종합 분석 도구")
    c.font      = Font(name="맑은 고딕", size=14, color=C["gray_d"], italic=True)
    c.alignment = align("center")

    # 구분선
    ws.merge_cells("B9:O9")
    c = ws.cell(row=9, column=1)
    ws.row_dimensions[9].height = 6
    for col in range(2, 16):
        ws.cell(row=9, column=col).fill = fill(C["blue"])

    # 이미지 자리
    ws.row_dimensions[10].height = 10
    placeholder(ws, 11, 2, 22, 15, "[ 프로그램 대표 화면 캡처 삽입 ]")

    # 문서 정보 표
    ws.row_dimensions[23].height = 8
    info = [
        ("문서명",   "핑감지 테스트기 v2 사용자 매뉴얼"),
        ("버전",     "v2.0"),
        ("작성일",   TODAY),
        ("작성자",   ""),
        ("부서",     ""),
        ("승인자",   ""),
    ]
    for i, (key, val) in enumerate(info, 24):
        ws.row_dimensions[i].height = 22
        ws.merge_cells(start_row=i, start_column=2,
                       end_row=i, end_column=4)
        ws.merge_cells(start_row=i, start_column=5,
                       end_row=i, end_column=14)
        k = ws.cell(row=i, column=2, value=key)
        k.font      = font(C["white"], 10, bold=True)
        k.fill      = fill(C["navy"])
        k.alignment = align("center")
        k.border    = thin_border()
        v = ws.cell(row=i, column=5, value=val)
        v.font      = font(size=10)
        v.fill      = fill(C["blue_lt"])
        v.alignment = align("left")
        v.border    = thin_border()

    # 로고 자리
    ws.merge_cells("B30:O32")
    c = ws.cell(row=30, column=2, value="[ 회사 로고 삽입 가능 영역 ]")
    c.font      = Font(name="맑은 고딕", size=10, color=C["gray_d"], italic=True)
    c.fill      = fill("F8F8F8")
    c.alignment = align("center")
    c.border    = thin_border()

    # 하단 장식
    ws.merge_cells("A33:P34")
    for col in range(1, 17):
        ws.cell(row=33, column=col).fill = fill(C["navy"])
    ws.row_dimensions[33].height = 8
    ws.row_dimensions[34].height = 8


# ═══════════════════════════════════════════════════════════════════════
# 시트2: 프로그램 개요
# ═══════════════════════════════════════════════════════════════════════
def build_overview(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "② 개요", SHEETS)

    for i, w in enumerate([3,18,18,18,18,18,18,3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "프로그램 개요", 8)

    # 목적
    set_section(ws, 5, "  1. 프로그램 목적", 8)
    ws.merge_cells("B6:H9")
    c = ws.cell(row=6, column=2,
                value="핑감지 테스트기 v2는 NC Agent 네트워크 연결 상태를 실시간으로 감시하여\n"
                      "장애 원인을 자동 분석하고 운영자가 신속하게 대응할 수 있도록 지원하는\n"
                      "NC Agent 네트워크 종합 분석 도구입니다.\n\n"
                      "현장 운영자, 생산기술, IT 담당자가 별도 프로그램 설치 없이 즉시 사용 가능합니다.")
    c.font      = font(size=10)
    c.fill      = fill(C["blue_lt"])
    c.alignment = align("left", wrap=True)
    c.border    = thin_border()
    for r in range(6, 10):
        ws.row_dimensions[r].height = 20

    # 주요 기능
    set_section(ws, 11, "  2. 주요 기능", 8)
    features = [
        ("Ping 감시",         "설비 IP 및 서버에 주기적 Ping 수행, 응답시간 기록"),
        ("NC Agent 감시",     "NCAgent.exe 프로세스 실행 여부, CPU/메모리 실시간 감시"),
        ("네트워크 어댑터",    "Link Up/Down, IP 변경, Gateway 변경 자동 감지"),
        ("Windows 이벤트",    "네트워크 관련 시스템 이벤트 로그 자동 수집"),
        ("장애 원인 추정",    "설비/서버 Ping 조합으로 원인 자동 분류 및 권장 조치 안내"),
        ("응답 누락 판단",    "연속 실패 횟수 기준 일시적 누락 vs 실제 장애 자동 판별"),
        ("Excel 보고서",      "매일 자정 일별 분석 보고서 자동 생성 (.xlsx)"),
        ("자동 스크린샷",     "장애 발생 시 현재 화면 자동 캡처 저장"),
        ("로그 자동 정리",    "설정 기간 경과 로그를 archive 폴더로 자동 이동"),
        ("실시간 그래프",     "최근 1시간 Ping 응답시간 추이 그래프 표시"),
    ]
    table_header(ws, 12, ["기능명", "설명"], 2, C["blue"])
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55
    for i, (fn, desc) in enumerate(features, 13):
        table_row(ws, i, [fn, desc], 2, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18

    # 적용 대상
    set_section(ws, 24, "  3. 적용 대상", 8)
    targets = [
        ("현장 운영자",     "장비 이상 발생 시 원인 확인 및 초동 조치"),
        ("생산기술 담당",   "설비 네트워크 장애 원인 분석 및 데이터 기록"),
        ("IT 담당자",       "네트워크 인프라 장애 원인 분석 및 이력 관리"),
    ]
    table_header(ws, 25, ["대상", "활용 목적"], 2, C["blue"])
    for i, (t, p) in enumerate(targets, 26):
        table_row(ws, i, [t, p], 2, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18

    # 시스템 구성
    set_section(ws, 30, "  4. 시스템 구성", 8)
    placeholder(ws, 31, 2, 40, 8,
                "[ 시스템 구성도 이미지 삽입 ]\n"
                "PC ↔ 스위치 ↔ 설비(192.168.0.101)\n"
                "PC ↔ 인터넷 ↔ hidc.cps.org")

    # 프로그램 파일 구성
    set_section(ws, 42, "  5. 프로그램 파일 구성", 8)
    files = [
        ("ping_monitor_gui.py",     "GUI 메인 프로그램 (Python 소스)"),
        ("핑감지 테스트기 실행.bat", "프로그램 실행기 (더블클릭 실행)"),
        ("build.bat",               "EXE 빌드 스크립트 (Python 없는 PC 배포용)"),
        ("requirements.txt",        "필요 패키지 목록 (psutil, openpyxl)"),
        ("config.json",             "사용자 설정 저장 (자동 생성)"),
        ("logs/",                   "로그 CSV 파일 저장 폴더 (자동 생성)"),
        ("logs/archive/",           "30일 경과 로그 자동 보관 폴더"),
        ("logs/screenshots/",       "장애 발생 시 자동 스크린샷 저장 폴더"),
    ]
    table_header(ws, 43, ["파일명", "설명"], 2, C["blue"])
    for i, (f, d) in enumerate(files, 44):
        table_row(ws, i, [f, d], 2, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18


# ═══════════════════════════════════════════════════════════════════════
# 시트3: 화면 구성 설명
# ═══════════════════════════════════════════════════════════════════════
def build_screen(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "③ 화면", SHEETS)
    for i in range(1, 17):
        ws.column_dimensions[get_column_letter(i)].width = 9

    set_title(ws, 3, "화면 구성 설명", 16)

    # 화면 캡처 자리
    set_section(ws, 5, "  전체 화면 구성", 16)
    placeholder(ws, 6, 1, 22, 16,
                "[ 프로그램 전체 화면 캡처 삽입 ]\n"
                "① 설정 영역  ② 실행 버튼  ③ 대시보드  ④ 로그 탭\n"
                "(실제 운영 시 캡처 이미지로 교체)")

    # 구역별 설명 표
    set_section(ws, 24, "  구역별 기능 설명", 16)
    headers = ["번호", "구분", "기능명", "설명", "비고"]
    table_header(ws, 25, headers, 1, C["navy"])
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 20

    areas = [
        ("①", "설정 영역",     "Ping 대상 설정",    "서버 및 설비 IP 등록, 모니터링 간격 설정, 로그 경로 설정", "더블클릭으로 수정 가능"),
        ("②", "설정 영역",     "실행 버튼",         "▶시작: 모니터링 시작 / ■중지: 모니터링 중지", "중지 시 보고서 자동 생성"),
        ("③", "대시보드",      "현재 시간",         "현재 날짜와 시각 표시 (1초마다 갱신)", "Power/Status LED 포함"),
        ("④", "대시보드",      "실시간 상태 카드",  "설비Ping / 서버Ping / NC Agent 상태 표시", "색상으로 OK/FAIL 구분"),
        ("⑤", "대시보드",      "네트워크 어댑터",   "사용 중인 랜카드 이름, IP, Gateway 표시", "Link Down 시 빨간색"),
        ("⑥", "대시보드",      "오늘 통계",         "총 Ping 횟수, 각 항목 실패 횟수 표시", "자정에 자동 초기화"),
        ("⑦", "대시보드",      "응답시간 그래프",   "최근 1시간 응답시간 추이 실시간 그래프", "마우스 오버 시 상세 표시"),
        ("⑧", "Ping 로그",     "전체 Ping 이력",    "모든 Ping 결과 시간순 기록 (CSV 저장)", "CSV로 저장 버튼 제공"),
        ("⑨", "프로세스",      "NC Agent 상태",     "NCAgent.exe PID, CPU, 메모리 5초마다 확인", "종료 감지 시 경고"),
        ("⑩", "네트워크",      "어댑터 이벤트",     "Link Up/Down, IP 변경, Gateway 변경 기록", "5초마다 확인"),
        ("⑪", "이벤트 로그",   "Windows 이벤트",    "네트워크 관련 시스템 이벤트 60초마다 수집", "관리자 권한 불필요"),
        ("⑫", "NC Agent 로그", "로그 자동 백업",    "NCAgent 설치 경로 자동 탐색, 로그 파일 변경 시 자동 백업", "30초마다 확인"),
        ("⑬", "장애 분석",     "원인 추정",         "현재 상태 조합으로 장애 원인 실시간 분석 및 권장 조치 표시", "6가지 패턴 분류"),
        ("⑭", "장애 분석",     "응답 누락 판단",    "연속 실패 횟수 기준 일시적 누락/실제 장애 자동 판별", "요약 박스 포함"),
    ]
    for i, row in enumerate(areas, 26):
        colors = None
        if "장애" in row[1] or "FAIL" in row[3]:
            colors = [None, C["red_l"], None, None, None]
        elif "대시보드" in row[1]:
            colors = [None, C["green_l"], None, None, None]
        table_row(ws, i, list(row), 1, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 22

    # 화면 흐름
    set_section(ws, 41, "  화면 흐름 (탭 구성)", 16)
    flow = [
        ("대시보드",      "실시간 현황 한눈에 보기"),
        ("Ping 로그",     "전체 Ping 기록 조회"),
        ("프로세스",      "NC Agent 프로세스 이력"),
        ("네트워크",      "어댑터 이벤트 이력"),
        ("이벤트 로그",   "Windows 시스템 이벤트"),
        ("NC Agent 로그", "Agent 로그 백업 이력"),
        ("장애 분석",     "자동 원인 분석 + 응답 누락 판단"),
    ]
    table_header(ws, 42, ["탭 이름", "주요 용도"], 1, C["blue"])
    for i, (t, d) in enumerate(flow, 43):
        table_row(ws, i, [t, d], 1, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18


# ═══════════════════════════════════════════════════════════════════════
# 시트4: 기능 설명
# ═══════════════════════════════════════════════════════════════════════
def build_features(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "④ 기능", SHEETS)
    for i, w in enumerate([6,20,20,25,25,14,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "기능 상세 설명", 7)

    features = [
        {
            "title": "1. Ping 감시 (설비 IP)",
            "rows": [
                ("기능",     "설정된 설비 IP로 주기적(기본 5초)으로 Ping 수행"),
                ("목적",     "설비와의 네트워크 연결 상태 실시간 확인"),
                ("입력",     "설비 IP 주소 (예: 192.168.0.101), Ping 간격(초)"),
                ("출력",     "응답 성공(OK)/실패(FAIL), 응답시간(ms) → ping_log.csv 저장"),
                ("활용",     "설비 통신 장애 시 최초 발생 시각 파악, 장애 패턴 분석"),
                ("비고",     "IP 추가/수정 가능, 개수 제한 없음, 설정은 자동 저장"),
            ]
        },
        {
            "title": "2. Ping 감시 (서버: hidc.cps.org)",
            "rows": [
                ("기능",     "NC Agent 서버(hidc.cps.org)로 주기적 Ping 수행"),
                ("목적",     "인터넷/WAN 연결 및 서버 응답 상태 확인"),
                ("입력",     "서버 도메인 또는 IP, Ping 간격"),
                ("출력",     "응답 성공/실패, 응답시간 → ping_log.csv 저장"),
                ("활용",     "인터넷 단절, 서버 점검 시간 확인"),
                ("비고",     "설비 Ping과 비교하여 장애 원인 자동 추정에 활용"),
            ]
        },
        {
            "title": "3. NC Agent 프로세스 감시",
            "rows": [
                ("기능",     "NCAgent.exe 프로세스 실행 여부를 5초마다 확인"),
                ("목적",     "NC Agent 비정상 종료 즉시 감지"),
                ("입력",     "없음 (자동 감지)"),
                ("출력",     "PID, CPU(%), 메모리(MB), 이벤트 → process_log.csv 저장"),
                ("활용",     "Agent 종료 시각 기록, 재시작 여부 확인"),
                ("비고",     "psutil 패키지 필요. 미설치 시 해당 기능 비활성화"),
            ]
        },
        {
            "title": "4. 네트워크 어댑터 감시",
            "rows": [
                ("기능",     "PC 랜카드 상태를 5초마다 확인"),
                ("목적",     "물리적 네트워크 단절(Link Down) 즉시 감지"),
                ("입력",     "없음 (자동 감지)"),
                ("출력",     "Link Up/Down, IP 변경, Gateway 변경 → network_log.csv 저장"),
                ("활용",     "케이블 단선, 스위치 이상 확인"),
                ("비고",     "psutil 패키지 필요"),
            ]
        },
        {
            "title": "5. Windows 이벤트 로그 수집",
            "rows": [
                ("기능",     "시스템 이벤트 로그에서 네트워크 관련 이벤트 60초마다 수집"),
                ("목적",     "OS 수준의 네트워크 오류 자동 기록"),
                ("수집 대상","Tcpip, DNS Client, NDIS, Kernel-PnP, Realtek, RTL8153 등"),
                ("출력",     "이벤트 시간, 소스, ID, 수준, 메시지 → event_log.csv 저장"),
                ("활용",     "드라이버 오류, DNS 실패 등 심층 원인 분석"),
                ("비고",     "PowerShell(Get-WinEvent) 사용, 관리자 권한 불필요"),
            ]
        },
        {
            "title": "6. 장애 원인 자동 추정",
            "rows": [
                ("기능",     "현재 Ping 상태 조합으로 장애 원인 자동 분류"),
                ("목적",     "운영자가 즉시 원인 파악 및 조치 가능하도록 안내"),
                ("판정 기준","설비+서버 동시 실패, 서버만 실패, 설비만 실패, Link Down 등 6가지"),
                ("출력",     "장애 분석 탭 원인 추정 패널에 실시간 표시"),
                ("활용",     "장애 발생 시 조치 우선순위 결정"),
                ("비고",     "Ping 루프 실행 시마다 자동 갱신"),
            ]
        },
        {
            "title": "7. Excel 일별 보고서 자동 생성",
            "rows": [
                ("기능",     "매일 자정에 전날 데이터 기반 Excel 보고서 자동 생성"),
                ("목적",     "일별 장애 현황 기록 및 보고"),
                ("포함 항목","총 Ping 횟수, FAIL 횟수, 최대/평균 응답시간, Agent 종료, 동시 장애 수"),
                ("출력",     "logs/daily_report_YYYYMMDD.xlsx"),
                ("활용",     "주간/월간 보고서 작성 기초 자료"),
                ("비고",     "■중지 또는 X 버튼 종료 시에도 즉시 생성"),
            ]
        },
    ]

    row_pos = 5
    for feat in features:
        set_section(ws, row_pos, f"  {feat['title']}", 7)
        row_pos += 1
        table_header(ws, row_pos, ["항목", "내용"], 2, C["blue"])
        row_pos += 1
        for i, (k, v) in enumerate(feat["rows"]):
            table_row(ws, row_pos, [k, v], 2, alt=(i % 2 == 0))
            ws.row_dimensions[row_pos].height = 20
            row_pos += 1
        row_pos += 1


# ═══════════════════════════════════════════════════════════════════════
# 시트5: 장애 분석 기능
# ═══════════════════════════════════════════════════════════════════════
def build_fault_analysis(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "⑤ 장애", SHEETS)
    for i, w in enumerate([6,20,22,28,28,6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "장애 분석 기능 상세", 6)

    # 화면 캡처
    set_section(ws, 5, "  장애 분석 탭 화면", 6)
    placeholder(ws, 6, 1, 18, 6, "[ 장애 분석 탭 화면 캡처 삽입 ]")

    # 판정 기준 표
    set_section(ws, 20, "  판정 기준 (응답 누락 분석)", 6)
    table_header(ws, 21, ["조건", "판정 결과", "추정 원인", "권장 조치"], 2, C["navy"])
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30

    criteria = [
        ("1~2회 연속 실패 후 복구",
         "일시적 응답 누락",
         "패킷 순간 손실 (네트워크 혼잡 등)",
         "지속 모니터링, 빈도 증가 시 점검",
         C["yell_l"]),
        ("3회 이상 연속 실패 (서버 정상)",
         "실제 네트워크 장애 가능성",
         "설비 전원 OFF, 케이블 단선, 스위치 포트 오류",
         "설비 전원·케이블·스위치 포트 점검",
         C["red_l"]),
        ("서버+설비 동시 실패",
         "공통 구간 장애 가능성",
         "PC 랜어댑터 또는 상위 공유기·스위치 문제",
         "공유기·스위치 재시작, PC 네트워크 설정 확인",
         C["purp_l"]),
        ("서버 정상 + 설비 실패",
         "설비망 구간 문제 가능성",
         "설비와 PC 사이 스위치·케이블 문제",
         "설비 IP 확인, 스위치 포트·케이블 점검",
         C["oran_l"]),
        ("Link Down 감지",
         "랜카드/케이블 연결 문제",
         "케이블 단선, 스위치 포트 이상",
         "랜 케이블 재연결, 스위치 전원 확인",
         C["red_l"]),
        ("Ping 정상 + NC Agent 종료",
         "NC Agent 프로그램 문제",
         "Agent 크래시, 업데이트 오류, 권한 문제",
         "Agent 로그 확인 후 재시작",
         C["yell_l"]),
    ]
    for i, (cond, verdict, cause, action, bg) in enumerate(criteria, 22):
        ws.row_dimensions[i].height = 24
        for j, val in enumerate([cond, verdict, cause, action], 2):
            c = ws.cell(row=i, column=j, value=val)
            c.fill      = fill(bg)
            c.font      = font(size=9)
            c.alignment = align("left", wrap=True)
            c.border    = thin_border()

    # 분석 알고리즘 순서도 (텍스트 기반)
    set_section(ws, 29, "  분석 알고리즘 순서도", 6)

    def flowbox(ws, row, col, text, bg, merge_cols=4):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + merge_cols - 1)
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(name="맑은 고딕", size=9, bold=True,
                           color=C["white"])
        c.fill      = fill(bg)
        c.alignment = align("center")
        c.border    = thick_border()
        ws.row_dimensions[row].height = 22
        return c

    def arrow(ws, row, col, merge_cols=4):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + merge_cols - 1)
        c = ws.cell(row=row, column=col, value="↓")
        c.font      = Font(name="맑은 고딕", size=11, color=C["gray_d"])
        c.alignment = align("center")
        ws.row_dimensions[row].height = 12

    flowbox(ws, 30, 2, "[ Ping 사이클 시작 ]", C["navy"])
    arrow(ws, 31, 2)
    flowbox(ws, 32, 2, "Ping 결과 수신 (OK / FAIL)", C["blue"])
    arrow(ws, 33, 2)
    flowbox(ws, 34, 2, "◇  실패 여부?", C["yell_h"])

    ws.row_dimensions[35].height = 18
    ws.merge_cells("B35:C35")
    c1 = ws.cell(row=35, column=2, value="  Yes → 연속 실패 카운트 증가")
    c1.font = font(C["red"], 9)
    c1.alignment = align("left")
    ws.merge_cells("D35:E35")
    c2 = ws.cell(row=35, column=4, value="No → 복구 감지?  →  분석 결과 기록")
    c2.font = font(C["green_h"], 9)
    c2.alignment = align("left")

    arrow(ws, 36, 2)
    flowbox(ws, 37, 2, "◇  연속 실패 횟수?", C["yell_h"])

    ws.row_dimensions[38].height = 20
    verdicts = [
        ("B", "1~2회 → 일시적 응답 누락 (노란색)", C["yell_l"]),
        ("C", "3회↑ + 서버 정상 → 설비망 장애 (빨간색)", C["red_l"]),
        ("D", "동시 실패 → 공통 구간 장애 (보라색)", C["purp_l"]),
        ("E", "서버 실패 포함 → 상위 네트워크 (빨간색)", C["red_l"]),
    ]
    for col_ltr, text, bg in verdicts:
        col_num = ord(col_ltr) - ord("A") + 1
        c = ws.cell(row=38, column=col_num, value=text)
        c.fill      = fill(bg)
        c.font      = font(size=8)
        c.alignment = align("center", wrap=True)
        c.border    = thin_border()

    arrow(ws, 39, 2)
    flowbox(ws, 40, 2, "분석 결과 테이블 기록 + 요약 박스 갱신", C["navy"])

    # 판정 문구 예시
    set_section(ws, 42, "  판정 결과 문구 예시", 6)
    examples = [
        ("일시적 응답 누락",
         '"1회 실패 후 즉시 복구되어 일시적 응답 누락으로 판단됩니다."',
         C["yell_l"]),
        ("실제 장애 가능성",
         '"3회 이상 연속 실패하여 실제 네트워크 장애 가능성이 있습니다."',
         C["red_l"]),
        ("공통 구간 장애",
         '"동시 실패로 스위치, 메인 PC 랜어댑터, NC Agent 공통 구간 확인이 필요합니다."',
         C["purp_l"]),
        ("설비망 문제",
         '"서버 Ping은 정상이나 설비 Ping만 실패하여 현장 설비망 점검이 필요합니다."',
         C["oran_l"]),
    ]
    table_header(ws, 43, ["판정 유형", "표시 문구"], 2, C["blue"])
    for i, (t, msg, bg) in enumerate(examples, 44):
        ws.row_dimensions[i].height = 24
        for j, val in enumerate([t, msg], 2):
            c = ws.cell(row=i, column=j, value=val)
            c.fill      = fill(bg)
            c.font      = font(size=9)
            c.alignment = align("left", wrap=True)
            c.border    = thin_border()


# ═══════════════════════════════════════════════════════════════════════
# 시트6: 운영 매뉴얼
# ═══════════════════════════════════════════════════════════════════════
def build_operations(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "⑥ 운영", SHEETS)
    for i, w in enumerate([4,20,18,40,20,8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "운영 매뉴얼 (단계별 절차)", 6)

    steps = [
        {
            "step": "STEP 1", "title": "프로그램 실행",
            "desc": "1. 탐색기에서 PingMonitor 폴더 열기\n"
                    "2. '핑감지 테스트기 실행.bat' 더블 클릭\n"
                    "3. Python 미설치 시 자동 설치 안내 메시지 확인\n"
                    "4. 프로그램 창이 열리면 정상 실행 완료\n\n"
                    "※ EXE 버전: PingMonitor.exe 더블 클릭",
            "img": "[ 실행 화면 캡처 삽입 ]",
            "color": C["blue"],
        },
        {
            "step": "STEP 2", "title": "Ping 대상 IP 등록",
            "desc": "1. IP 설정 영역에서 기존 항목 확인\n"
                    "2. '+ 추가' 버튼 클릭 → 이름과 IP 입력\n"
                    "3. 기존 항목 수정: 행 더블 클릭 → 수정 팝업\n"
                    "4. 불필요한 항목: 행 선택 후 '- 삭제' 클릭\n"
                    "5. 간격(초) 설정: 기본값 5초 (1~300초)",
            "img": "[ IP 설정 화면 캡처 삽입 ]",
            "color": C["blue"],
        },
        {
            "step": "STEP 3", "title": "모니터링 시작",
            "desc": "1. 로그 저장 경로 확인 (기본: 프로그램 폴더\\logs\\)\n"
                    "2. '▶ 시작' 버튼 클릭\n"
                    "3. Power LED → 녹색, Status LED → 노란색으로 변경\n"
                    "4. 상태 표시줄에 '모니터링 시작' 메시지 확인\n"
                    "5. 대시보드에서 실시간 상태 확인",
            "img": "[ 시작 후 대시보드 캡처 삽입 ]",
            "color": C["green"],
        },
        {
            "step": "STEP 4", "title": "Ping 로그 확인",
            "desc": "1. 상단 탭에서 'Ping 로그' 클릭\n"
                    "2. OK(초록) / FAIL(빨간) 항목 확인\n"
                    "3. 특정 구간 분석: 시간순 정렬 확인\n"
                    "4. CSV 저장: 하단 '📥 CSV로 저장' 버튼\n"
                    "5. 원본 CSV: logs\\ping_log.csv",
            "img": "[ Ping 로그 탭 캡처 삽입 ]",
            "color": C["blue"],
        },
        {
            "step": "STEP 5", "title": "장애 분석 확인",
            "desc": "1. '장애 분석' 탭 클릭\n"
                    "2. 상단 요약 박스에서 오늘 현황 확인\n"
                    "3. '원인 추정 & 장애 이력' 탭: 실시간 원인 분석 확인\n"
                    "4. '응답 누락 분석' 탭: 일시적 누락 vs 실제 장애 확인\n"
                    "5. 색상 기준: 노란색=일시적, 빨간색=실제 장애, 보라색=공통",
            "img": "[ 장애 분석 탭 캡처 삽입 ]",
            "color": C["yell_h"],
        },
        {
            "step": "STEP 6", "title": "장애 발생 시 조치",
            "desc": "1. Status LED 빨간색 → 장애 발생 확인\n"
                    "2. 대시보드 실시간 상태 확인\n"
                    "3. 장애 분석 탭 → 원인 추정 패널에서 원인 확인\n"
                    "4. 권장 조치 항목 수행\n"
                    "5. logs\\screenshots\\에서 자동 저장된 스크린샷 확인\n"
                    "6. 조치 후 OK 복구 확인 → Status LED 노란색 복귀",
            "img": "[ 장애 발생 화면 캡처 삽입 ]",
            "color": C["red"],
        },
        {
            "step": "STEP 7", "title": "보고서 확인 및 모니터링 종료",
            "desc": "1. '■ 중지' 버튼 클릭 → 당일 보고서 자동 생성\n"
                    "2. 보고서 위치: logs\\daily_report_YYYYMMDD.xlsx\n"
                    "3. Excel로 보고서 열기 → 일별 통계 확인\n"
                    "4. X 버튼 클릭 → 종료 확인 팝업 → '종료' 클릭\n"
                    "5. 설정(IP, 간격 등)은 자동 저장됨",
            "img": "[ 종료 확인 팝업 캡처 삽입 ]",
            "color": C["navy"],
        },
    ]

    row_pos = 5
    for step in steps:
        # 스텝 헤더
        ws.row_dimensions[row_pos].height = 28
        ws.merge_cells(start_row=row_pos, start_column=1,
                       end_row=row_pos, end_column=6)
        c = ws.cell(row=row_pos, column=1,
                    value=f"  {step['step']}  {step['title']}")
        c.font      = Font(name="맑은 고딕", size=12, bold=True,
                           color=C["white"])
        c.fill      = fill(step["color"])
        c.alignment = align("left")
        row_pos += 1

        # 설명 + 이미지 placeholder
        ws.merge_cells(start_row=row_pos, start_column=2,
                       end_row=row_pos + 6, end_column=4)
        dc = ws.cell(row=row_pos, column=2, value=step["desc"])
        dc.font      = font(size=9)
        dc.fill      = fill(C["blue_lt"])
        dc.alignment = align("left", wrap=True)
        dc.border    = thin_border()

        ws.merge_cells(start_row=row_pos, start_column=5,
                       end_row=row_pos + 6, end_column=6)
        ic = ws.cell(row=row_pos, column=5, value=step["img"])
        ic.font      = Font(name="맑은 고딕", size=9, color=C["gray_d"],
                            italic=True)
        ic.fill      = fill("F2F2F2")
        ic.alignment = align("center")
        ic.border    = thin_border()

        for r in range(row_pos, row_pos + 7):
            ws.row_dimensions[r].height = 18

        row_pos += 8


# ═══════════════════════════════════════════════════════════════════════
# 시트7: 장애 유형별 조치
# ═══════════════════════════════════════════════════════════════════════
def build_fault_types(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "⑦ 조치", SHEETS)
    for i, w in enumerate([6,22,24,28,28,6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "장애 유형별 조치 방법", 6)

    types = [
        {
            "type": "일시적 응답 누락",
            "color": C["yell_l"],
            "hcolor": C["yell_h"],
            "rows": [
                ("원인",    "순간 네트워크 혼잡, 패킷 손실, 일시적 처리 지연"),
                ("증상",    "1~2회 FAIL 후 즉시 OK 복구, 응답 누락 판단 탭: 노란색"),
                ("확인",    "Ping 로그에서 해당 시간대 FAIL 횟수 및 빈도 확인"),
                ("조치",    "빈도가 낮으면(1~2회/일) 정상 허용 범위\n"
                            "빈도 증가 시 → 스위치 포트 점검 또는 케이블 교체 검토"),
                ("기준",    "1~2회 연속 실패: 일시적 누락\n3회 이상: 실제 장애 의심"),
            ]
        },
        {
            "type": "설비 통신 장애 (서버 정상)",
            "color": C["oran_l"],
            "hcolor": C["orange"],
            "rows": [
                ("원인",    "설비 전원 OFF, 랜 케이블 단선, 스위치 포트 이상, 설비 IP 변경"),
                ("증상",    "설비 Ping FAIL, 서버 Ping OK, 응답 누락 판단: 빨간색"),
                ("확인",    "① 설비 전원 상태 확인\n② 설비↔스위치 간 랜 케이블 연결 확인\n③ 스위치 포트 Link LED 확인"),
                ("조치",    "① 설비 전원 켜기\n② 케이블 재연결 또는 교체\n③ 스위치 포트 교체\n④ 설비 IP 설정 확인"),
                ("기준",    "3회 이상 연속 실패 + 서버 Ping 정상"),
            ]
        },
        {
            "type": "서버 통신 장애 (설비 정상)",
            "color": C["red_l"],
            "hcolor": C["red"],
            "rows": [
                ("원인",    "인터넷(WAN) 단절, 서버 점검 또는 장애, DNS 오류"),
                ("증상",    "서버 Ping FAIL, 설비 Ping OK, NC Agent 연결 불가"),
                ("확인",    "① 웹 브라우저에서 인터넷 연결 확인\n② hidc.cps.org 서버 운영 상태 문의\n③ DNS 설정 확인 (cmd → nslookup hidc.cps.org)"),
                ("조치",    "① 공유기 재시작\n② 인터넷 서비스 업체(ISP) 문의\n③ 서버 운영팀 문의"),
                ("기준",    "서버 Ping 3회 이상 연속 실패 + 설비 Ping 정상"),
            ]
        },
        {
            "type": "공통 구간 장애 (동시 실패)",
            "color": C["purp_l"],
            "hcolor": C["purple"],
            "rows": [
                ("원인",    "PC 랜어댑터 이상, 상위 스위치/공유기 장애, 정전"),
                ("증상",    "설비·서버 Ping 동시 FAIL, Link Down 이벤트 발생"),
                ("확인",    "① 이벤트 로그 탭에서 Link Down 이벤트 확인\n② 네트워크 탭에서 Link 상태 확인\n③ 공유기·스위치 전원 및 포트 LED 확인"),
                ("조치",    "① 공유기·스위치 전원 재시작\n② PC 랜어댑터 드라이버 재설치\n③ 네트워크 케이블 교체"),
                ("기준",    "서버·설비 Ping 동시 실패 또는 Link Down"),
            ]
        },
        {
            "type": "NC Agent 종료",
            "color": C["yell_l"],
            "hcolor": C["yell_h"],
            "rows": [
                ("원인",    "Agent 프로그램 오류(크래시), Windows 업데이트, 권한 문제"),
                ("증상",    "프로세스 탭에서 '종료 감지' 이벤트, Ping은 정상"),
                ("확인",    "① NC Agent 로그 탭에서 로그 내용 확인\n② Windows 이벤트 뷰어에서 응용 프로그램 오류 확인\n③ 작업 관리자에서 NCAgent.exe 확인"),
                ("조치",    "① NC Agent 재시작 (시작 메뉴 또는 서비스)\n② Agent 로그 확인 후 개발사 문의\n③ 지속 발생 시 재설치 검토"),
                ("기준",    "Ping 정상 + NCAgent.exe 프로세스 미감지"),
            ]
        },
        {
            "type": "랜어댑터 Link Down",
            "color": C["red_l"],
            "hcolor": C["red"],
            "rows": [
                ("원인",    "랜 케이블 물리적 단선, 스위치 포트 이상, 어댑터 드라이버 오류"),
                ("증상",    "네트워크 탭에서 'Link Down' 이벤트, 이벤트 로그에서 NDIS 오류"),
                ("확인",    "① 랜 케이블 연결 상태 확인\n② PC 랜포트 및 스위치 포트 LED 확인\n③ 이벤트 로그에서 NDIS/네트워크 어댑터 오류 확인"),
                ("조치",    "① 케이블 재연결 또는 교체\n② 다른 스위치 포트로 변경\n③ 네트워크 어댑터 드라이버 재설치\n④ 랜카드 교체 검토"),
                ("기준",    "네트워크 탭 Link Down + 이벤트 로그 NDIS 오류"),
            ]
        },
    ]

    row_pos = 5
    for fault in types:
        ws.row_dimensions[row_pos].height = 26
        ws.merge_cells(start_row=row_pos, start_column=1,
                       end_row=row_pos, end_column=6)
        c = ws.cell(row=row_pos, column=1,
                    value=f"  ■  {fault['type']}")
        c.font      = Font(name="맑은 고딕", size=11, bold=True,
                           color=C["white"])
        c.fill      = fill(fault["hcolor"])
        c.alignment = align("left")
        row_pos += 1

        table_header(ws, row_pos, ["항목", "내용"], 2, C["blue"])
        row_pos += 1

        for i, (k, v) in enumerate(fault["rows"]):
            ws.row_dimensions[row_pos].height = 28
            kc = ws.cell(row=row_pos, column=2, value=k)
            kc.font      = font(C["navy"], 9, bold=True)
            kc.fill      = fill(fault["color"])
            kc.alignment = align("center")
            kc.border    = thin_border()
            vc = ws.cell(row=row_pos, column=3, value=v)
            ws.merge_cells(start_row=row_pos, start_column=3,
                           end_row=row_pos, end_column=6)
            vc.font      = font(size=9)
            vc.fill      = fill(C["white"] if i % 2 == 0 else C["blue_lt"])
            vc.alignment = align("left", wrap=True)
            vc.border    = thin_border()
            row_pos += 1

        row_pos += 1


# ═══════════════════════════════════════════════════════════════════════
# 시트8: FAQ
# ═══════════════════════════════════════════════════════════════════════
def build_faq(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "⑧ FAQ", SHEETS)
    for i, w in enumerate([4,14,44,6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "자주 묻는 질문 (FAQ)", 4)

    faqs = [
        {
            "q": "Q1. 설비 Ping은 정상인데 hidc.cps.org(서버)만 OFF가 발생합니다.",
            "a": "A. 설비 Ping이 정상이므로 사내 네트워크는 정상입니다.\n"
                 "   원인: 인터넷(WAN) 연결 이상 또는 hidc.cps.org 서버 장애 가능성이 높습니다.\n"
                 "   조치: ① 웹 브라우저로 인터넷 연결 확인\n"
                 "         ② 공유기 WAN 포트 상태 확인\n"
                 "         ③ 서버 운영팀 문의 (서버 점검 일정 확인)"
        },
        {
            "q": "Q2. 1초 OFF가 발생했는데 이것도 장애인가요?",
            "a": "A. 1~2회(5~10초) 연속 실패는 '일시적 응답 누락'으로 분류됩니다.\n"
                 "   실제 장애로 판정되지 않으며 정상 운영 범위 내일 수 있습니다.\n"
                 "   조치: 빈도를 모니터링하여 하루 3회 이상 발생 시 스위치·케이블 점검 권장"
        },
        {
            "q": "Q3. NC Agent가 종료되었다고 표시됩니다.",
            "a": "A. NCAgent.exe 프로세스가 비정상 종료된 것입니다.\n"
                 "   조치: ① NC Agent 재시작\n"
                 "         ② 'NC Agent 로그' 탭에서 자동 백업된 로그 파일 내용 확인\n"
                 "         ③ Windows 이벤트 로그 탭에서 오류 이벤트 확인\n"
                 "         ④ 반복 발생 시 개발사(HI-CPS) 문의"
        },
        {
            "q": "Q4. 설비 10대가 동시에 OFF되었습니다.",
            "a": "A. 여러 설비가 동시에 실패하면 '공통 구간 장애'로 분류됩니다.\n"
                 "   원인: 설비 공통 구간(메인 스위치, 공유기, 랜어댑터) 문제 가능성이 높습니다.\n"
                 "   조치: ① 공통 스위치 및 허브 전원 상태 확인\n"
                 "         ② 이벤트 로그 탭에서 Link Down 이벤트 확인\n"
                 "         ③ PC 랜어댑터 상태 확인"
        },
        {
            "q": "Q5. Ping 응답시간이 갑자기 증가했습니다.",
            "a": "A. 그래프에서 주황색 구간으로 표시됩니다 (이전 대비 50% 이상 급증).\n"
                 "   원인: 네트워크 혼잡, 스위치 성능 저하, 바이러스/보안 소프트웨어 간섭\n"
                 "   조치: ① 응답시간 추이 그래프에서 지속 여부 확인\n"
                 "         ② 200ms 이상 지속 시 네트워크 점검 필요\n"
                 "         ③ 바이러스 검사 수행"
        },
        {
            "q": "Q6. 로그 파일이 너무 커졌습니다.",
            "a": "A. 로그 자동 정리 기능을 사용하세요.\n"
                 "   방법: 설정 영역 > 로그 보관 일수 설정(기본 30일) > '지금 정리' 버튼\n"
                 "   자동: 모니터링 중 24시간마다 자동으로 archive 폴더로 이동됨\n"
                 "   위치: logs\\archive\\ping_log_YYYYMM.csv"
        },
        {
            "q": "Q7. 다른 PC에서도 사용하려면 어떻게 해야 하나요?",
            "a": "A. 두 가지 방법이 있습니다:\n"
                 "   방법 1 (Python 있는 PC): PingMonitor 폴더 전체 복사 → 실행.bat 더블클릭\n"
                 "   방법 2 (Python 없는 PC): 현재 PC에서 build.bat 실행하여 PingMonitor.exe 생성\n"
                 "                          → 해당 EXE 파일만 복사하면 어디서나 실행 가능"
        },
        {
            "q": "Q8. Excel 보고서가 생성되지 않습니다.",
            "a": "A. openpyxl 패키지가 필요합니다.\n"
                 "   설치: 명령 프롬프트에서 'pip install openpyxl' 입력 후 재실행\n"
                 "   또는: '핑감지 테스트기 실행.bat'을 실행하면 자동으로 설치됩니다."
        },
        {
            "q": "Q9. 스크린샷이 저장되지 않습니다.",
            "a": "A. 스크린샷은 PowerShell(.NET Framework)을 사용합니다.\n"
                 "   확인: ① 로그 경로 폴더에 쓰기 권한 확인\n"
                 "         ② Windows Defender 또는 보안 소프트웨어 차단 여부 확인\n"
                 "         ③ 저장 경로: logs\\screenshots\\"
        },
        {
            "q": "Q10. 모니터링 간격은 어떻게 변경하나요?",
            "a": "A. 설정 영역 > '간격' 스핀박스에서 숫자 변경 (1~300초)\n"
                 "   기본값: 5초. 변경 사항은 자동 저장됩니다.\n"
                 "   주의: 1초로 설정 시 로그 파일이 빠르게 증가할 수 있습니다."
        },
    ]

    row_pos = 5
    for faq in faqs:
        ws.row_dimensions[row_pos].height = 24
        ws.merge_cells(start_row=row_pos, start_column=2,
                       end_row=row_pos, end_column=4)
        qc = ws.cell(row=row_pos, column=2, value=faq["q"])
        qc.font      = Font(name="맑은 고딕", size=10, bold=True,
                            color=C["white"])
        qc.fill      = fill(C["navy"])
        qc.alignment = align("left")
        qc.border    = thin_border()
        row_pos += 1

        lines = faq["a"].count("\n") + 3
        ws.row_dimensions[row_pos].height = max(18 * lines, 60)
        ws.merge_cells(start_row=row_pos, start_column=2,
                       end_row=row_pos, end_column=4)
        ac = ws.cell(row=row_pos, column=2, value=faq["a"])
        ac.font      = font(size=9)
        ac.fill      = fill(C["blue_lt"])
        ac.alignment = align("left", wrap=True)
        ac.border    = thin_border()
        row_pos += 2


# ═══════════════════════════════════════════════════════════════════════
# 시트9: 유지보수 이력
# ═══════════════════════════════════════════════════════════════════════
def build_history(ws):
    a4_landscape(ws)
    add_nav_bar(ws, "⑨ 이력", SHEETS)
    for i, w in enumerate([4,14,10,50,18,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    set_title(ws, 3, "유지보수 이력", 6)

    set_section(ws, 5, "  프로그램 버전 이력", 6)
    table_header(ws, 6, ["날짜", "버전", "변경 내용", "작성자", "비고"], 2, C["navy"])
    history = [
        ("2026-06-19", "v2.0",
         "NC Agent 프로세스 감시, 네트워크 어댑터 감시,\n"
         "Windows 이벤트 로그, 장애 원인 추정, 응답 누락 판단,\n"
         "Excel 보고서, 실시간 그래프, 로그 자동 정리 추가",
         "", "초기 배포"),
        ("2026-06-19", "v2.0.1", "LED 배경색 오류 수정 (ttk 호환)", "", "버그픽스"),
        ("2026-06-19", "v2.0.2", "이식성 개선: 경로 하드코딩 제거, config.json 분리", "", "개선"),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
    ]
    for i, row in enumerate(history, 7):
        ws.row_dimensions[i].height = 36 if i <= 9 else 22
        for j, val in enumerate(row, 2):
            c = ws.cell(row=i, column=j, value=val)
            c.fill      = fill(C["blue_lt"] if i % 2 == 0 else C["white"])
            c.font      = font(size=9)
            c.alignment = align("left", wrap=True)
            c.border    = thin_border()

    # 향후 개발 계획
    set_section(ws, 17, "  향후 개발 계획", 6)
    plans = [
        ("1", "알림 기능", "장애 발생 시 소리 알림, Windows 팝업 알림, 이메일 발송"),
        ("2", "NC Agent 자동 재시작", "종료 감지 시 자동 재시작 (야간 무인 운용 대응)"),
        ("3", "임계값 경고", "응답시간 100ms 초과 시 WARN 상태로 선제 경고"),
        ("4", "주간/월간 보고서", "일별 보고서를 취합한 주간·월간 추세 분석 보고서"),
        ("5", "다중 PC 지원", "여러 PC에서 로그 집계 및 중앙 모니터링"),
    ]
    table_header(ws, 18, ["#", "기능명", "내용"], 2, C["blue"])
    for i, row in enumerate(plans, 19):
        table_row(ws, i, list(row), 2, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 20

    # 문의처
    set_section(ws, 25, "  문의처", 6)
    contacts = [
        ("GitHub", "https://github.com/wqsaxzedc79-glitch/PingMonitor"),
        ("이메일", ""),
        ("담당자", ""),
    ]
    for i, (k, v) in enumerate(contacts, 26):
        ws.row_dimensions[i].height = 20
        ws.merge_cells(start_row=i, start_column=2,
                       end_row=i, end_column=2)
        ws.merge_cells(start_row=i, start_column=3,
                       end_row=i, end_column=6)
        kc = ws.cell(row=i, column=2, value=k)
        kc.font      = font(C["white"], 9, bold=True)
        kc.fill      = fill(C["navy"])
        kc.alignment = align("center")
        kc.border    = thin_border()
        vc = ws.cell(row=i, column=3, value=v)
        vc.font      = font(size=9)
        vc.fill      = fill(C["blue_lt"])
        vc.alignment = align("left")
        vc.border    = thin_border()


# ═══════════════════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    builders = [
        ("표지",             build_cover),
        ("프로그램 개요",    build_overview),
        ("화면 구성 설명",   build_screen),
        ("기능 설명",        build_features),
        ("장애 분석 기능",   build_fault_analysis),
        ("운영 매뉴얼",      build_operations),
        ("장애 유형별 조치", build_fault_types),
        ("FAQ",              build_faq),
        ("유지보수 이력",    build_history),
    ]

    for i, (name, builder) in enumerate(builders):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)

        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale     = 85
        builder(ws)
        print(f"  [OK] {name}")

    # 탭 색상
    tab_colors = [
        "1F3864", "2E75B6", "375623", "4472C4",
        "C55A11", "843C0C", "7F6000", "595959", "375623"
    ]
    for ws, color in zip(wb.worksheets, tab_colors):
        ws.sheet_properties.tabColor = color

    wb.save(OUT)
    print(f"\n[완료] 매뉴얼 생성: {OUT}")
    return OUT


if __name__ == "__main__":
    main()
